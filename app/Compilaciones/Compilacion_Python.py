# ==========================================
# COMPILACION AUTOMATICA DE PROYECTO DJANGO
# ARCHIVO: Compilacion_Python.py
# FECHA: 2026-01-26 15:12:18
# ==========================================

--------------------------------------------------
# ARCHIVO: admin.py
# RUTA: C:\Users\juanl\Git\PROYECTOS\UrbanStore\UStore\app\admin.py
--------------------------------------------------

from django.contrib import admin
from .models import Marca, Tronco, Piernas, Zapatos, Complemento, Contacto, TrabajaConNosotros, SeSocio, Envio, Tienda

class TroncoAdmin(admin.ModelAdmin):
    list_display = ["modelo", "precio", "cantidad", "nuevo", "marca"]
    list_editable = ["precio", "cantidad", "nuevo", "marca"]
    search_fields = ["modelo"]
    list_filter = ["marca", "nuevo"]
    list_per_page = 10

class PiernasAdmin(admin.ModelAdmin):
    list_display = ["modelo", "precio", "cantidad", "nuevo", "marca"]
    list_editable = ["precio", "cantidad", "nuevo", "marca"]
    search_fields = ["modelo"]
    list_filter = ["marca", "nuevo"]
    list_per_page = 10

class ZapatosAdmin(admin.ModelAdmin):
    list_display = ["modelo", "precio", "cantidad", "nuevo", "marca"]
    list_editable = ["precio", "cantidad", "nuevo", "marca"]
    search_fields = ["modelo"]
    list_filter = ["marca", "nuevo"]
    list_per_page = 10

class ComplementoAdmin(admin.ModelAdmin):
    list_display = ["modelo", "precio", "cantidad", "nuevo", "marca", "tipo"]
    list_editable = ["precio", "cantidad", "nuevo", "marca", "tipo"]
    search_fields = ["modelo"]
    list_filter = ["marca", "nuevo", "tipo"]
    list_per_page = 10

class ContactoAdmin(admin.ModelAdmin):
    list_display = ['nombre', 'email', 'fecha', 'leido']
    list_filter = ['leido', 'fecha']
    search_fields = ['nombre', 'email', 'mensaje']
    list_editable = ['leido']
    date_hierarchy = 'fecha'

class TrabajaConNosotrosAdmin(admin.ModelAdmin):
    list_display = ['nombre', 'email', 'puesto', 'fecha', 'leido']
    list_filter = ['puesto', 'leido', 'fecha']
    search_fields = ['nombre', 'email', 'mensaje']
    list_editable = ['leido']
    date_hierarchy = 'fecha'
    readonly_fields = ['fecha']

class SeSocioAdmin(admin.ModelAdmin):
    list_display = ['nombre_marca', 'nombre', 'email', 'rubro', 'fecha', 'contacto']
    list_filter = ['rubro', 'contacto', 'fecha']
    search_fields = ['nombre_marca', 'nombre', 'email']
    list_editable = ['contacto']
    date_hierarchy = 'fecha'

class EnvioAdmin(admin.ModelAdmin):
    list_display = ['nombre', 'ciudad', 'region', 'total_pagar', 'fecha', 'marcas_compradas']
    list_filter = ['region', 'fecha']
    search_fields = ['nombre', 'email']

class TiendaAdmin(admin.ModelAdmin):
    list_display = ['nombre_tienda', 'categorias']
    search_fields = ['nombre_tienda', 'categorias']
    list_per_page = 10

admin.site.register(SeSocio, SeSocioAdmin)
admin.site.register(TrabajaConNosotros, TrabajaConNosotrosAdmin)
admin.site.register(Contacto, ContactoAdmin)
admin.site.register(Marca)
admin.site.register(Tronco, TroncoAdmin)
admin.site.register(Piernas, PiernasAdmin)
admin.site.register(Zapatos, ZapatosAdmin)
admin.site.register(Complemento, ComplementoAdmin)
admin.site.register(Envio, EnvioAdmin)
admin.site.register(Tienda, TiendaAdmin)

--------------------------------------------------
# ARCHIVO: apps.py
# RUTA: C:\Users\juanl\Git\PROYECTOS\UrbanStore\UStore\app\apps.py
--------------------------------------------------

from django.apps import AppConfig


class AppConfig(AppConfig):
    default_auto_field = 'django.db.models.BigAutoField'
    name = 'app'
    verbose_name = "25UrbanStore"


--------------------------------------------------
# ARCHIVO: forms.py
# RUTA: C:\Users\juanl\Git\PROYECTOS\UrbanStore\UStore\app\forms.py
--------------------------------------------------

# forms.py
from django import forms
from .models import Marca, Tronco, Piernas, Zapatos, Complemento, Contacto, GENERO_CHOICES, COLOR_CHOICES, SUBCATEGORIA_TRONCO, SUBCATEGORIA_PIERNAS, SUBCATEGORIA_ZAPATOS, TrabajaConNosotros, SeSocio, Envio
from django.contrib.auth.forms import UserCreationForm
import datetime

class MarcaForm(forms.ModelForm):
    class Meta:
        model = Marca
        fields = ['nombre']
        widgets = {
            'nombre': forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': 'Ingrese el nombre de la marca'
            }),
        }

class ProductoForm(forms.Form):
    TIPO_PRODUCTO = [
        ('Tronco', 'Torso'),
        ('Piernas', 'Pantalones'),
        ('Zapatos', 'Calzado'),
        ('Complemento', 'Accesorios'),
    ]
    
    tipo = forms.ChoiceField(
        choices=TIPO_PRODUCTO,
        widget=forms.Select(attrs={'class': 'form-control'}),
        label='Tipo de Producto'
    )
    
    modelo = forms.CharField(
        max_length=50,
        required=False,
        widget=forms.TextInput(attrs={'class': 'form-control'}),
        label='Modelo'
    )
    
    precio = forms.IntegerField(
        widget=forms.NumberInput(attrs={'class': 'form-control'}),
        label='Precio'
    )
    
    descripcion = forms.CharField(
        widget=forms.Textarea(attrs={'class': 'form-control', 'rows': 3}),
        label='Descripción'
    )
    
    nuevo = forms.BooleanField(
        required=False,
        widget=forms.CheckboxInput(attrs={'class': 'form-check-input'}),
        label='¿Es nuevo?'
    )
    
    marca = forms.ModelChoiceField(
        queryset=Marca.objects.all(),
        widget=forms.Select(attrs={'class': 'form-control'}),
        label='Marca'
    )
    
    imagen = forms.ImageField(
        required=False,
        widget=forms.FileInput(attrs={'class': 'form-control'}),
        label='Imagen'
    )
    
    tipo_accesorio = forms.ChoiceField(
        choices=Complemento.TIPO_ACCESORIO,
        required=False,
        widget=forms.Select(attrs={'class': 'form-control'}),
        label='Tipo de Accesorio'
    )

    cantidad = forms.IntegerField(
        min_value=0,
        widget=forms.NumberInput(attrs={'class': 'form-control'}),
        label='Cantidad en stock',
        initial=0
    )
    genero = forms.ChoiceField(
        choices=GENERO_CHOICES,
        widget=forms.Select(attrs={'class': 'form-control'}),
        label='Género',
        initial='UNI'
    )
    
    color = forms.ChoiceField(
        choices=COLOR_CHOICES,
        widget=forms.Select(attrs={'class': 'form-control'}),
        label='Color',
        initial='NEG'
    )
    
    # Campos para subcategorías (condicionales)
    subcategoria_tronco = forms.ChoiceField(
        choices=SUBCATEGORIA_TRONCO,
        required=False,
        widget=forms.Select(attrs={'class': 'form-control'}),
        label='Subcategoría'
    )
    
    subcategoria_piernas = forms.ChoiceField(
        choices=SUBCATEGORIA_PIERNAS,
        required=False,
        widget=forms.Select(attrs={'class': 'form-control'}),
        label='Subcategoría'
    )
    
    subcategoria_zapatos = forms.ChoiceField(
        choices=SUBCATEGORIA_ZAPATOS,
        required=False,
        widget=forms.Select(attrs={'class': 'form-control'}),
        label='Subcategoría'
    )

class ContactoForm(forms.ModelForm):
    class Meta:
        model = Contacto
        fields = ['nombre', 'email', 'telefono', 'mensaje']
        widgets = {
            'nombre': forms.TextInput(attrs={'class': 'form-control'}),
            'email': forms.EmailInput(attrs={'class': 'form-control'}), 
            'telefono': forms.TextInput(attrs={'class': 'form-control'}),
            'mensaje': forms.Textarea(attrs={'class': 'form-control', 'rows': 4}),
        }

class TrabajaConNosotrosForm(forms.ModelForm):
    class Meta:
        model = TrabajaConNosotros
        fields = ['nombre', 'email', 'telefono', 'puesto', 'mensaje', 'cv']
        widgets = {
            'nombre': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Nombre completo'}),
            'email': forms.EmailInput(attrs={'class': 'form-control', 'placeholder': 'Correo electrónico'}),
            'telefono': forms.TextInput(attrs={'class': 'form-control', 'placeholder': 'Teléfono'}),
            'puesto': forms.Select(attrs={'class': 'form-select'}),
            'mensaje': forms.Textarea(attrs={'class': 'form-control', 'rows': 4, 'placeholder': 'Cuéntanos por qué quieres trabajar con nosotros...'}),
            'cv': forms.FileInput(attrs={'class': 'form-control', 'accept': '.pdf,.doc,.docx'})
        }
        labels = {
            'cv': 'Curriculum Vitae (PDF o Word)'
        }

class SeSocioForm(forms.ModelForm):
    class Meta:
        model = SeSocio
        fields = ['nombre', 'email', 'rut_comercial', 'nombre_marca', 'rubro', 'descripcion']
        widgets = {
            'nombre': forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': 'Nombre completo'
            }),
            'email': forms.EmailInput(attrs={
                'class': 'form-control',
                'placeholder': 'Correo electrónico'
            }),
            'rut_comercial': forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': 'RUT comercial (ej: 12.345.678-9)'
            }),
            'nombre_marca': forms.TextInput(attrs={
                'class': 'form-control',
                'placeholder': 'Nombre de tu marca'
            }),
            'rubro': forms.Select(attrs={
                'class': 'form-select'
            }),
            'descripcion': forms.Textarea(attrs={
                'class': 'form-control',
                'rows': 4,
                'placeholder': 'Describe los productos que vendes, tu experiencia, etc...'
            }),
        }
        labels = {
            'nombre_marca': 'Nombre de la Marca',
            'rut_comercial': 'RUT Comercial',
        }
        
class EnvioForm(forms.ModelForm):
    class Meta:
        model = Envio
        fields = ['nombre', 'direccion', 'ciudad', 'region', 'codigo_postal', 'telefono', 'email', 'courier']
        widgets = {
            'nombre': forms.TextInput(attrs={'class': 'form-control'}),
            'direccion': forms.TextInput(attrs={'class': 'form-control'}),
            'ciudad': forms.TextInput(attrs={'class': 'form-control'}),
            'region': forms.Select(attrs={'class': 'form-control', 'id': 'region-select'}),
            'courier': forms.Select(attrs={'class': 'form-control', 'id': 'courier-select'}),
            'codigo_postal': forms.TextInput(attrs={'class': 'form-control'}),
            'telefono': forms.TextInput(attrs={'class': 'form-control'}),
            'email': forms.EmailInput(attrs={'class': 'form-control'}),
        }

class TiendaForm(forms.Form):
    CATEGORIAS_CHOICES = [
        ('Torso', 'Torso'),
        ('Pantalones', 'Pantalones'),
        ('Calzado', 'Calzado'),
        ('Accesorios', 'Accesorios'),
    ]

    nombre_tienda = forms.CharField(
        max_length=100,
        required=True,
        widget=forms.TextInput(attrs={
            'class': 'urban-input',
            'placeholder': 'Ej: Zona Norte, Outlet Maipú...',
            'id': 'nombre_tienda'
        }),
        label='NOMBRE DEL SPOT'
    )

    categorias = forms.MultipleChoiceField(
        choices=CATEGORIAS_CHOICES,
        widget=forms.CheckboxSelectMultiple,
        required=True,
        label='¿QUÉ MUEVES AQUÍ?'
    )

--------------------------------------------------
# ARCHIVO: models.py
# RUTA: C:\Users\juanl\Git\PROYECTOS\UrbanStore\UStore\app\models.py
--------------------------------------------------

from django.db import models
from django.contrib.auth.models import AbstractUser, User
from django.db import models
import datetime

GENERO_CHOICES = [
    ('HOM', 'Hombre'),
    ('MUJ', 'Mujer'),
    ('UNI', 'Unisex'),
]

COLOR_CHOICES = [
    ('NEG', 'Negro'),
    ('BLN', 'Blanco'),
    ('GRS', 'Gris'),
    ('AZM', 'Azul marino'),
    ('BEI', 'Beige'),
    ('ROJ', 'Rojo'),
    ('VDO', 'Verde oliva'),
    ('MAR', 'Marrón'),
    ('AZC', 'Azul claro'),
    ('ROS', 'Rosado'),
]

# Subcategorías para cada tipo de producto
SUBCATEGORIA_TRONCO = [
    ('POL', 'Poleras'),
    ('PLN', 'Poleron'),
    ('CAM', 'Camisas'),
]

SUBCATEGORIA_PIERNAS = [
    ('JEA', 'Jeans'),
    ('BUS', 'Buso'),
    ('SHT', 'Short'),
    ('FAL', 'Falda'),
    ('VES', 'Vestidos'),
]

SUBCATEGORIA_ZAPATOS = [
    ('ZAP', 'Zapatillas'),
    ('BOT', 'Botos'),
    ('SAN', 'Sandalias'),
    ('MOC', 'Mocasines'),
    ('PAN', 'Pantuflas'),
]

class Marca(models.Model):
    nombre = models.CharField(max_length=50)
    # --- NUEVOS CAMPOS AGREGADOS ---
    usuario = models.ForeignKey(User, on_delete=models.SET_NULL, null=True, blank=True, verbose_name="Creado por")
    tienda = models.ForeignKey('Tienda', on_delete=models.SET_NULL, null=True, blank=True, verbose_name="Tienda asociada")
    # -------------------------------

    def __str__(self):
        return self.nombre

class Tronco(models.Model):  # Antes: Torso
    modelo = models.CharField(max_length=50, default='', blank=True)
    precio = models.IntegerField()
    descripcion = models.TextField()
    nuevo = models.BooleanField()
    marca = models.ForeignKey(Marca, on_delete=models.PROTECT)
    imagen = models.ImageField(upload_to="productos", null=True)
    cantidad = models.PositiveIntegerField(default=0) 
    genero = models.CharField(max_length=3, choices=GENERO_CHOICES, default='UNI')
    color = models.CharField(max_length=3, choices=COLOR_CHOICES, default='NEG')
    subcategoria = models.CharField(max_length=3, choices=SUBCATEGORIA_TRONCO, default='POL')
    def __str__(self):
        return self.modelo

class Piernas(models.Model):  # Antes: Pantalones
    modelo = models.CharField(max_length=50, default='', blank=True)
    precio = models.IntegerField()
    descripcion = models.TextField()
    nuevo = models.BooleanField()
    marca = models.ForeignKey(Marca, on_delete=models.PROTECT)
    imagen = models.ImageField(upload_to="productos", null=True)
    cantidad = models.PositiveIntegerField(default=0)
    genero = models.CharField(max_length=3, choices=GENERO_CHOICES, default='UNI')
    color = models.CharField(max_length=3, choices=COLOR_CHOICES, default='NEG')
    subcategoria = models.CharField(max_length=3, choices=SUBCATEGORIA_PIERNAS, default='JEA')
    def __str__(self):
        return self.modelo

class Zapatos(models.Model):  # Antes: Calzado
    modelo = models.CharField(max_length=50, default='', blank=True)
    precio = models.IntegerField()
    descripcion = models.TextField()
    nuevo = models.BooleanField()
    marca = models.ForeignKey(Marca, on_delete=models.PROTECT)
    imagen = models.ImageField(upload_to="productos", null=True)
    cantidad = models.PositiveIntegerField(default=0)
    genero = models.CharField(max_length=3, choices=GENERO_CHOICES, default='UNI')
    color = models.CharField(max_length=3, choices=COLOR_CHOICES, default='NEG')
    subcategoria = models.CharField(max_length=3, choices=SUBCATEGORIA_ZAPATOS, default='ZAP')
    def __str__(self):
        return self.modelo

class Complemento(models.Model):  # Antes: Accesorios
    TIPO_ACCESORIO = [
        ('REL', 'Reloj'),
        ('BUF', 'Bufanda'),
        ('CIN', 'Cinturón'),
        ('COL', 'Collar'),
        ('GOR', 'Gorra'),
        ('PUL', 'Pulseras'),
    ]
    
    modelo = models.CharField(max_length=50, default='', blank=True)
    precio = models.IntegerField()
    descripcion = models.TextField()
    nuevo = models.BooleanField()
    marca = models.ForeignKey(Marca, on_delete=models.PROTECT)
    tipo = models.CharField(max_length=3, choices=TIPO_ACCESORIO, default='REL')
    imagen = models.ImageField(upload_to="productos", null=True)
    cantidad = models.PositiveIntegerField(default=0)
    genero = models.CharField(max_length=3, choices=GENERO_CHOICES, default='UNI')
    color = models.CharField(max_length=3, choices=COLOR_CHOICES, default='NEG')
    
    def __str__(self):
        return self.modelo

class Contacto(models.Model):
    nombre = models.CharField(max_length=100)
    email = models.EmailField()
    telefono = models.CharField(max_length=20, blank=True, null=True)
    mensaje = models.TextField()
    fecha = models.DateTimeField(auto_now_add=True)
    leido = models.BooleanField(default=False)

    def __str__(self):
        return f"Mensaje de {self.nombre} - {self.fecha.strftime('%d/%m/%Y %H:%M')}"
    
class TrabajaConNosotros(models.Model):
    PUESTO_CHOICES = [
        ('VEN', 'Vendedor'),
        ('LID', 'Líder de Equipo'),
    ]
    
    nombre = models.CharField(max_length=100)
    email = models.EmailField()
    telefono = models.CharField(max_length=20)
    puesto = models.CharField(max_length=3, choices=PUESTO_CHOICES)
    mensaje = models.TextField()
    cv = models.FileField(upload_to="cvs/")
    fecha = models.DateTimeField(auto_now_add=True)
    leido = models.BooleanField(default=False)

    def __str__(self):
        return f"Solicitud de {self.nombre} para {self.get_puesto_display()} - {self.fecha.strftime('%d/%m/%Y')}"

class SeSocio(models.Model):
    RUBRO_CHOICES = [
        ('ROP', 'Ropa'),
        ('CAL', 'Calzado'),
        ('ACC', 'Accesorios'),
        ('DEP', 'Deportes'),
        ('OTR', 'Otros'),
    ]
    
    nombre = models.CharField(max_length=100)
    email = models.EmailField()
    rut_comercial = models.CharField(max_length=20)
    nombre_marca = models.CharField(max_length=100)
    rubro = models.CharField(max_length=3, choices=RUBRO_CHOICES)
    descripcion = models.TextField()
    fecha = models.DateTimeField(auto_now_add=True)
    contacto = models.BooleanField(default=False)

    def __str__(self):
        return f"Solicitud de {self.nombre_marca} por {self.nombre}"
    
class Envio(models.Model):
    REGION_CHOICES = [
        ('RM', 'Región Metropolitana de Santiago'),
        ('I', 'Región de Tarapacá'),
        ('II', 'Región de Antofagasta'),
        ('III', 'Región de Atacama'),
        ('IV', 'Región de Coquimbo'),
        ('V', 'Región de Valparaíso'),
        ('VI', 'Región de O’Higgins'),
        ('VII', 'Región del Maule'),
        ('VIII', 'Región del Biobío'),
        ('IX', 'Región de La Araucanía'),
        ('XIV', 'Región de Los Ríos'),
        ('X', 'Región de Los Lagos'),
        ('XI', 'Región de Aysén'),
        ('XII', 'Región de Magallanes'),
        ('XV', 'Región de Arica y Parinacota'),
        ('XVI', 'Región de Ñuble'),
    ]
    
    COSTOS_ENVIO = {
        'RM': 3000,
        'I': 13000,
        'II': 14000,
        'III': 11000,
        'IV': 8000,
        'V': 4000,
        'VI': 5000,
        'VII': 6000,
        'VIII': 7000,
        'IX': 9000,
        'XIV': 10000,
        'X': 12000,
        'XI': 15000,
        'XII': 16000,
        'XV': 14500,
        'XVI': 7000,
    }

    COURIER_CHOICES = [
        ('CHILEXPRESS', 'Chilexpress'),
        ('STARKEN', 'Starken'),
        ('CORREOS_CHILE', 'Correos de Chile'),
        ('BLUE_EXPRESS', 'Blue Express'),
    ]

    COURIER_COSTS = {
        'CHILEXPRESS': 5000,
        'STARKEN': 4500,
        'CORREOS_CHILE': 4000,
        'BLUE_EXPRESS': 4800,
    }

    courier = models.CharField(max_length=20, choices=COURIER_CHOICES, default='CHILEXPRESS')
    marcas_compradas = models.CharField(max_length=255, blank=True, null=True, verbose_name="Marcas compradas")
    usuario = models.ForeignKey(User, on_delete=models.CASCADE, null=True, blank=True)
    nombre = models.CharField(max_length=100)
    direccion = models.CharField(max_length=200)
    ciudad = models.CharField(max_length=100)
    region = models.CharField(max_length=4, choices=REGION_CHOICES)
    codigo_postal = models.CharField(max_length=10)
    telefono = models.CharField(max_length=20)
    email = models.EmailField()
    monto_carrito = models.IntegerField()
    costo_envio = models.IntegerField()
    total_pagar = models.IntegerField()
    fecha = models.DateTimeField(auto_now_add=True)
    procesado = models.BooleanField(default=False) 

    def __str__(self):
        return f"Envío a {self.nombre} - {self.get_region_display()}"
    
    def calcular_costo_envio(self):
        costo_region = self.COSTOS_ENVIO.get(self.region, 0)
        costo_courier = self.COURIER_COSTS.get(self.courier, 0)
        return costo_region + costo_courier
    
    def get_region_cost(self):
        return self.COSTOS_ENVIO.get(self.region, 0)
    
    def get_courier_cost(self):
        return self.COURIER_COSTS.get(self.courier, 0)

class Tienda(models.Model):
    nombre_tienda = models.CharField(max_length=100, verbose_name="Nombre del Spot")
    categorias = models.CharField(max_length=255, verbose_name="Categorías que vende")

    def __str__(self):
        return self.nombre_tienda
    
class PreferenciasUsuario(models.Model):
    user = models.OneToOneField(User, on_delete=models.CASCADE, related_name="preferencias")
    recibir_notificaciones = models.BooleanField(default=False)

    def __str__(self):
        return f"Preferencias de {self.user.username}"

--------------------------------------------------
# ARCHIVO: tests.py
# RUTA: C:\Users\juanl\Git\PROYECTOS\UrbanStore\UStore\app\tests.py
--------------------------------------------------

from django.test import TestCase

# Create your tests here.


--------------------------------------------------
# ARCHIVO: transbank_utils.py
# RUTA: C:\Users\juanl\Git\PROYECTOS\UrbanStore\UStore\app\transbank_utils.py
--------------------------------------------------

# transbank_utils.py
import requests
from django.conf import settings

def header_request_transbank():
    return {
        "Authorization": "Token",
        "Tbk-Api-Key-Id": settings.TRANSBANK_API_KEY_ID,
        "Tbk-Api-Key-Secret": settings.TRANSBANK_API_KEY_SECRET,
        "Content-Type": "application/json",
        "Access-Control-Allow-Origin": "*",
        'Referrer-Policy': 'origin-when-cross-origin',
    }

def transbank_create(data):
    url = "https://webpay3gint.transbank.cl/rswebpaytransaction/api/webpay/v1.2/transactions"
    headers = header_request_transbank()
    response = requests.post(url, json=data, headers=headers)
    return response

def transbank_commit(tokenws):
    url = f"https://webpay3gint.transbank.cl/rswebpaytransaction/api/webpay/v1.2/transactions/{tokenws}"
    headers = header_request_transbank()
    response = requests.put(url, headers=headers)
    return response

def transbank_reverse_or_cancel(tokenws, amount):
    data = {"amount": amount}
    url = f"https://webpay3gint.transbank.cl/rswebpaytransaction/api/webpay/v1.2/transactions/{tokenws}/refunds"
    headers = header_request_transbank()
    response = requests.post(url, json=data, headers=headers)
    return response

--------------------------------------------------
# ARCHIVO: urls.py
# RUTA: C:\Users\juanl\Git\PROYECTOS\UrbanStore\UStore\app\urls.py
--------------------------------------------------

from django.urls import path
from . import views

urlpatterns = [
    path('', views.index, name="index"),
    path('Pantalones/', views.Pantalones, name="Pantalones"),
    path('Torso/', views.Torso, name="Torso"),
    path('Calzado/', views.Calzado, name="Calzado"),
    path('Accesorios/', views.Accesorios, name="Accesorios"),
    path('Contacto/', views.Contacto, name="Contacto"),
    path('Ventas/carrito/', views.carrito, name="carrito"),
    path('Ventas/Envio/', views.envio_view, name="Envio"),
    path('Ventas/Pago/', views.pago_view, name="Pago"),
    path('Ventas/commit_pay/', views.commit_pay, name="commit_pay"),
    path('Ventas/pago_error/', views.pago_error, name="pago_error"),
    path('Ventas/send_pay/', views.send_pay, name="send_pay"),
    path('producto/Administracion/', views.Administracion, name="Administracion"),
    path('producto/agregar_producto/', views.agregar_producto, name="agregar_producto"),
    path('producto/agrega_marca/', views.agrega_marca, name="agrega_marca"),
    path('producto/agregar_tienda/', views.agregar_tienda, name="agregar_tienda"),
    path('producto/modificar/', views.modificar, name="modificar"),
    path('producto/eliminar/', views.eliminar, name="eliminar"),
    path('Usuarios/IniciarSesion/', views.IniciarSesion, name="IniciarSesion"),
    path('Usuarios/Registrarse/', views.Registrarse, name="Registrarse"),
    path('Usuarios/Admin_User/', views.Admin_User, name="Admin_User"),
    path('Usuarios/FormularioEnvio/', views.FormularioEnvio, name="FormularioEnvio"),
    path('Footer/QuienesSomos/', views.QuienesSomos, name="QuienesSomos"),
    path('Footer/SeSocio/', views.SeSocio, name="SeSocio"),
    path('Footer/TrabajaConNosotros/', views.TrabajaConNosotros, name="TrabajaConNosotros"),
    path('Footer/TerminosCondiciones/', views.TerminosCondiciones, name="TerminosCondiciones"),
    
    # Rutas para API AJAX
    path('api/cart_details/', views.get_cart_data, name="api_cart_details"),
    path('update_cart/', views.update_cart, name="update_cart"), # Asegurando que esta ruta existe para los botones
]

--------------------------------------------------
# ARCHIVO: views.py
# RUTA: C:\Users\juanl\Git\PROYECTOS\UrbanStore\UStore\app\views.py
--------------------------------------------------

from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from .forms import MarcaForm, ProductoForm, ContactoForm, TrabajaConNosotrosForm, SeSocioForm, EnvioForm, TiendaForm
from .models import (
    GENERO_CHOICES, 
    COLOR_CHOICES, 
    SUBCATEGORIA_TRONCO, 
    SUBCATEGORIA_PIERNAS, 
    SUBCATEGORIA_ZAPATOS,
    Marca, Tronco, Piernas, Zapatos, Complemento, Envio as EnvioModel, Tienda
)
from .models import PreferenciasUsuario
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.forms import UserCreationForm, UserChangeForm
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
import json
from django.contrib.auth.decorators import user_passes_test
from django.conf import settings
from .transbank_utils import transbank_create, transbank_commit, transbank_reverse_or_cancel
import datetime as dt
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Create your views here.
TEMPLATE_DIRS = (
    'os.path.join(BASE_DIR, "templates")'
)

def index(request):
    # Obtener el producto más caro de cada categoría
    tronco_mas_caro = Tronco.objects.order_by('-precio').first()
    piernas_mas_caro = Piernas.objects.order_by('-precio').first()
    zapatos_mas_caro = Zapatos.objects.order_by('-precio').first()
    complemento_mas_caro = Complemento.objects.order_by('-precio').first()
    
    return render(request, "index.html", {
        'tronco_mas_caro': tronco_mas_caro,
        'piernas_mas_caro': piernas_mas_caro,
        'zapatos_mas_caro': zapatos_mas_caro,
        'complemento_mas_caro': complemento_mas_caro
    })

def envio_view(request):
    print("Accediendo a envio_view")  # Depuración
    cart = request.session.get('cart', {})
    print(f"Carrito en sesión: {cart}")  # Depuración
    cart = request.session.get('cart', {})
    if not cart:
        messages.error(request, 'Tu carrito está vacío. Agrega productos antes de proceder al envío.')
        return redirect('carrito')
    total_carrito = request.session.get('total_carrito', 0)
    
    marcas = set()
    for tipo, items in cart.items():
        for id_producto in items.keys():
            try:
                modelo = None
                if tipo == 'Tronco':
                    modelo = Tronco.objects.get(id=id_producto)
                elif tipo == 'Piernas':
                    modelo = Piernas.objects.get(id=id_producto)
                elif tipo == 'Zapatos':
                    modelo = Zapatos.objects.get(id=id_producto)
                elif tipo == 'Complemento':
                    modelo = Complemento.objects.get(id=id_producto)
                
                if modelo and modelo.marca:
                    marcas.add(modelo.marca.nombre)
            except:
                continue
    
    marcas_str = ", ".join(marcas) if marcas else ""

    if request.method == 'POST':
        form = EnvioForm(request.POST)
        if form.is_valid():
            envio = form.save(commit=False)
            envio.monto_carrito = total_carrito
            envio.costo_envio = envio.calcular_costo_envio()
            envio.total_pagar = envio.monto_carrito + envio.costo_envio
            envio.marcas_compradas = marcas_str  # Asignar las marcas
            
            if request.user.is_authenticated:
                envio.usuario = request.user
                
            envio.save()
            request.session['envio_id'] = envio.id
            return redirect('Pago')
        else:
            messages.error(request, 'Por Favor, corrige los errores en el formulario')
    else:
        form = EnvioForm()
    
    return render(request, "Ventas/Envio.html", {
        'form': form,
        'total_carrito': total_carrito
    })

def pago_view(request):
    envio_id = request.session.get('envio_id')
    if not envio_id:
        return redirect('carrito')
    
    try:
        envio = EnvioModel.objects.get(id=envio_id)
    except EnvioModel.DoesNotExist:
        return redirect('carrito')
    
    if request.method == 'POST':
        # Procesar el pago y descontar del stock
        cart = request.session.get('cart', {})
        
        for product_type, products in cart.items():
            model = None
            if product_type == 'Tronco':
                model = Tronco
            elif product_type == 'Piernas':
                model = Piernas
            elif product_type == 'Zapatos':
                model = Zapatos
            elif product_type == 'Complemento':
                model = Complemento
            
            if model:
                for product_id, quantity in products.items():
                    try:
                        product = model.objects.get(id=product_id)
                        if product.cantidad >= quantity:
                            product.cantidad -= quantity
                            product.save()
                        else:
                            messages.error(request, f'No hay suficiente stock para {product.modelo}')
                            return redirect('carrito')
                    except model.DoesNotExist:
                        continue
            # Crear transacción en Transbank
            buy_order = f"USTORE_{envio_id}"
            session_id = request.session.session_key
            amount = envio.total_pagar
            return_url = settings.TRANSBANK_RETURN_URL

            body = {
                "buy_order": buy_order,
                "session_id": session_id,
                "amount": amount,
                "return_url": return_url
            }

            response = transbank_create(body)
            if response.status_code == 200:
                transbank_data = response.json()
                return render(request, 'Ventas/send_pay.html', {
                    'transbank': transbank_data,
                    'amount': amount
                })
            else:
                return render(request, 'Ventas/pago_error.html', {
                    'error': 'Error al crear la transacción en Transbank'
                })
        
        # Limpiar el carrito después del pago
        if 'cart' in request.session:
            del request.session['cart']
        if 'total_carrito' in request.session:
            del request.session['total_carrito']
        if 'envio_id' in request.session:
            del request.session['envio_id']
        
        messages.success(request, '¡Pago realizado con éxito! Tu pedido ha sido procesado.')
        return redirect('index')
    
    return render(request, "Ventas/Pago.html", {
        'envio': envio
    })

@csrf_exempt
def commit_pay_view(request):
    transaction_detail = None
    try:
        tokenws = request.GET.get('token_ws') or request.POST.get('token_ws')
        
        if tokenws:
            response = transbank_commit(tokenws)
            
            if response.status_code == 200:
                response_data = response.json()
                status = response_data.get('status')
                response_code = response_data.get('response_code')
                
                if status == 'AUTHORIZED' and response_code == 0:
                    # Pago exitoso: descontar stock
                    cart = request.session.get('cart', {})
                    for product_type, products in cart.items():
                        model = None
                        if product_type == 'Tronco':
                            model = Tronco
                        elif product_type == 'Piernas':
                            model = Piernas
                        elif product_type == 'Zapatos':
                            model = Zapatos
                        elif product_type == 'Complemento':
                            model = Complemento
                        
                        if model:
                            for product_id, quantity in products.items():
                                try:
                                    product = model.objects.get(id=product_id)
                                    if product.cantidad >= quantity:
                                        product.cantidad -= quantity
                                        product.save()
                                    else:
                                        # Reversar pago si no hay stock suficiente
                                        amount = response_data.get('amount')
                                        reverse_response = transbank_reverse_or_cancel(tokenws, amount)
                                        return render(request, 'Ventas/pago_error.html', {
                                            'error': f'No hay suficiente stock para {product.modelo}'
                                        })
                                except model.DoesNotExist:
                                    pass
                    
                    # Limpiar carrito después del pago exitoso
                    if 'cart' in request.session:
                        del request.session['cart']
                    if 'total_carrito' in request.session:
                        del request.session['total_carrito']
                    if 'envio_id' in request.session:
                        del request.session['envio_id']
                    
                    # Preparar datos para mostrar
                    state = 'ACEPTADO'
                    pay_type = 'Tarjeta de Crédito' if response_data.get('payment_type_code') == 'VC' else 'Tarjeta de Débito'
                    amount = int(response_data.get('amount', 0))
                    amount_formatted = f'{amount:,.0f}'.replace(',', '.')
                    transaction_date = dt.datetime.strptime(response_data['transaction_date'], '%Y-%m-%dT%H:%M:%S.%fZ')
                    transaction_date = transaction_date.strftime('%d-%m-%Y %H:%M:%S')
                    
                    transaction_detail = {
                        'card_number': response_data['card_detail']['card_number'],
                        'transaction_date': transaction_date,
                        'state': state,
                        'pay_type': pay_type,
                        'amount': amount_formatted,
                        'authorization_code': response_data['authorization_code'],
                        'buy_order': response_data['buy_order'],
                    }
                else:
                    # Pago rechazado
                    state = 'RECHAZADO'
                    pay_type = 'Tarjeta de Crédito' if response_data.get('payment_type_code') == 'VC' else 'Tarjeta de Débito'
                    amount = int(response_data.get('amount', 0))
                    amount_formatted = f'{amount:,.0f}'.replace(',', '.')
                    transaction_date = dt.datetime.strptime(response_data['transaction_date'], '%Y-%m-%dT%H:%M:%S.%fZ')
                    transaction_date = transaction_date.strftime('%d-%m-%Y %H:%M:%S')
                    
                    # Intentar reversar el pago
                    reverse_response = transbank_reverse_or_cancel(tokenws, amount)
                    
                    transaction_detail = {
                        'card_number': response_data['card_detail']['card_number'],
                        'transaction_date': transaction_date,
                        'state': state,
                        'pay_type': pay_type,
                        'amount': amount_formatted,
                        'authorization_code': response_data['authorization_code'],
                        'buy_order': response_data['buy_order'],
                    }
        
        return render(request, 'Ventas/commit_pay.html', {
            'transaction_detail': transaction_detail
        })
    
    except Exception as e:
        return render(request, 'Ventas/pago_error.html', {
            'error': f'Error en el proceso de pago: {str(e)}'
        })

@login_required
def carrito(request):
    cart = request.session.get('cart', {})
    items = []
    total = 0
    
    for product_type, products in cart.items():
        model = None
        if product_type == 'Tronco':
            model = Tronco
        elif product_type == 'Piernas':
            model = Piernas
        elif product_type == 'Zapatos':
            model = Zapatos
        elif product_type == 'Complemento':
            model = Complemento
        
        if model:
            for product_id, quantity in products.items():
                try:
                    product = model.objects.get(id=product_id)
                    
                    # MOVER LA VERIFICACIÓN DE STOCK DENTRO DEL BLOQUE
                    if quantity > product.cantidad:
                        messages.warning(
                            request, 
                            f'No hay suficiente stock de {product.modelo}. Disponible: {product.cantidad}'
                        )
                    
                    subtotal = product.precio * quantity
                    total += subtotal
                    items.append({
                        'producto': product,
                        'tipo': product_type,
                        'cantidad': quantity,
                        'subtotal': subtotal
                    })
                except model.DoesNotExist:
                    continue
                    
    # Guardar total en sesión para usar en envío
    request.session['total_carrito'] = total

    return render(request, 'Ventas/carrito.html', {
        'items': items,
        'total': total
    })

def get_cart_data(request):
    """
    Vista API para devolver los datos del carrito en JSON
    para el sidebar lateral.
    """
    cart = request.session.get('cart', {})
    items_data = []
    total = 0
    
    for product_type, products in cart.items():
        model = None
        if product_type == 'Tronco':
            model = Tronco
        elif product_type == 'Piernas':
            model = Piernas
        elif product_type == 'Zapatos':
            model = Zapatos
        elif product_type == 'Complemento':
            model = Complemento
        
        if model:
            for product_id, quantity in products.items():
                try:
                    product = model.objects.get(id=product_id)
                    subtotal = product.precio * quantity
                    total += subtotal
                    
                    # Preparar imagen
                    img_url = ""
                    if product.imagen:
                        img_url = product.imagen.url
                    
                    items_data.append({
                        'id': product.id,
                        'modelo': product.modelo,
                        'precio': product.precio,
                        'cantidad': quantity,
                        'subtotal': subtotal,
                        'imagen': img_url,
                        'tipo': product_type,
                        'marca': product.marca.nombre if product.marca else ""
                    })
                except model.DoesNotExist:
                    continue
    
    return JsonResponse({
        'success': True,
        'items': items_data,
        'total': total,
        'total_items': len(items_data)
    })

@require_POST
@login_required
def update_cart(request):
    try:
        data = json.loads(request.body)
        product_id = str(data.get('product_id'))
        product_type = data.get('product_type')
        action = data.get('action')
        
        # Obtener el modelo del producto
        model = None
        if product_type == 'Tronco':
            model = Tronco
        elif product_type == 'Piernas':
            model = Piernas
        elif product_type == 'Zapatos':
            model = Zapatos
        elif product_type == 'Complemento':
            model = Complemento
        
        if not model:
            return JsonResponse({'success': False, 'message': 'Tipo de producto no válido'})
        
        # Obtener el producto
        try:
            product = model.objects.get(id=product_id)
        except model.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Producto no encontrado'})
        
        if not request.session.get('cart'):
            request.session['cart'] = {}
        
        cart = request.session['cart']
        
        if action == 'add':
            # Verificar stock disponible
            current_quantity = cart.get(product_type, {}).get(product_id, 0)
            if current_quantity + 1 > product.cantidad:
                return JsonResponse({
                    'success': False, 
                    'message': f'No hay suficiente stock. Disponible: {product.cantidad}'
                })
            
            if product_type not in cart:
                cart[product_type] = {}
            
            cart[product_type][product_id] = current_quantity + 1
            request.session.modified = True
            
        elif action == 'update':
            quantity = int(data.get('quantity', 1))
            if product_type in cart and product_id in cart[product_type]:
                cart[product_type][product_id] = quantity
                request.session.modified = True
            else:
                return JsonResponse({'success': False, 'message': 'Producto no encontrado en el carrito'})
        
        elif action == 'remove':
            if product_type in cart and product_id in cart[product_type]:
                del cart[product_type][product_id]
                # Eliminar la categoría si queda vacía
                if not cart[product_type]:
                    del cart[product_type]
                request.session.modified = True
            else:
                return JsonResponse({'success': False, 'message': 'Producto no encontrado en el carrito'})
        else:
            return JsonResponse({'success': False, 'message': 'Acción no válida'})
        
        # Calcular el total de items en el carrito
        total_items = sum(sum(products.values()) for products in cart.values())
        
        return JsonResponse({
            'success': True,
            'cart_count': total_items
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

def Torso(request):
    # Obtener parámetros de filtrado
    precio_min = request.GET.get('precio_min')
    precio_max = request.GET.get('precio_max')
    color = request.GET.get('color')
    genero = request.GET.get('genero')
    orden = request.GET.get('orden', 'reciente')  # Por defecto: ordenar por más reciente
    
    # Base query
    productos = Tronco.objects.all()
    
    # Aplicar filtros
    if precio_min:
        productos = productos.filter(precio__gte=precio_min)
    if precio_max:
        productos = productos.filter(precio__lte=precio_max)
    if color:
        productos = productos.filter(color=color)
    if genero:
        productos = productos.filter(genero=genero)
    
    # Aplicar ordenamiento
    if orden == 'precio_asc':
        productos = productos.order_by('precio')
    elif orden == 'precio_desc':
        productos = productos.order_by('-precio')
    else:  # 'reciente' o cualquier otro
        productos = productos.order_by('-id')  # Orden por defecto: más reciente primero
    
    # Paginación
    paginator = Paginator(productos, 12)  # 12 productos por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'COLOR_CHOICES': COLOR_CHOICES,
        'GENERO_CHOICES': GENERO_CHOICES,
        'current_filters': {
            'precio_min': precio_min,
            'precio_max': precio_max,
            'color': color,
            'genero': genero,
            'orden': orden
        },
        'template_name': 'Torso'  # Nueva variable
    }
    return render(request, "Torso.html", context)

def Pantalones(request):
    # Obtener parámetros de filtrado
    precio_min = request.GET.get('precio_min')
    precio_max = request.GET.get('precio_max')
    color = request.GET.get('color')
    genero = request.GET.get('genero')
    orden = request.GET.get('orden', 'reciente')
    
    # Base query
    productos = Piernas.objects.all()
    
    # Aplicar filtros
    if precio_min:
        productos = productos.filter(precio__gte=precio_min)
    if precio_max:
        productos = productos.filter(precio__lte=precio_max)
    if color:
        productos = productos.filter(color=color)
    if genero:
        productos = productos.filter(genero=genero)
    
    # Aplicar ordenamiento
    if orden == 'precio_asc':
        productos = productos.order_by('precio')
    elif orden == 'precio_desc':
        productos = productos.order_by('-precio')
    else:
        productos = productos.order_by('-id')
    
    # Paginación
    paginator = Paginator(productos, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'COLOR_CHOICES': COLOR_CHOICES,
        'GENERO_CHOICES': GENERO_CHOICES,
        'current_filters': {
            'precio_min': precio_min,
            'precio_max': precio_max,
            'color': color,
            'genero': genero,
            'orden': orden
        },
        'template_name': 'Pantalones' 
    }
    return render(request, "Pantalones.html", context)

def Calzado(request):
    # Obtener parámetros de filtrado
    precio_min = request.GET.get('precio_min')
    precio_max = request.GET.get('precio_max')
    color = request.GET.get('color')
    genero = request.GET.get('genero')
    orden = request.GET.get('orden', 'reciente')
    
    # Base query
    productos = Zapatos.objects.all()
    
    # Aplicar filtros
    if precio_min:
        productos = productos.filter(precio__gte=precio_min)
    if precio_max:
        productos = productos.filter(precio__lte=precio_max)
    if color:
        productos = productos.filter(color=color)
    if genero:
        productos = productos.filter(genero=genero)
    
    # Aplicar ordenamiento
    if orden == 'precio_asc':
        productos = productos.order_by('precio')
    elif orden == 'precio_desc':
        productos = productos.order_by('-precio')
    else:
        productos = productos.order_by('-id')
    
    # Paginación
    paginator = Paginator(productos, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'COLOR_CHOICES': COLOR_CHOICES,
        'GENERO_CHOICES': GENERO_CHOICES,
        'current_filters': {
            'precio_min': precio_min,
            'precio_max': precio_max,
            'color': color,
            'genero': genero,
            'orden': orden
        },
        'template_name': 'Calzado'
    }
    return render(request, "Calzado.html", context)

def Accesorios(request):
    # Obtener parámetros de filtrado
    precio_min = request.GET.get('precio_min')
    precio_max = request.GET.get('precio_max')
    color = request.GET.get('color')
    genero = request.GET.get('genero')
    orden = request.GET.get('orden', 'reciente')
    
    # Base query
    productos = Complemento.objects.all()
    
    # Aplicar filtros
    if precio_min:
        productos = productos.filter(precio__gte=precio_min)
    if precio_max:
        productos = productos.filter(precio__lte=precio_max)
    if color:
        productos = productos.filter(color=color)
    if genero:
        productos = productos.filter(genero=genero)
    
    # Aplicar ordenamiento
    if orden == 'precio_asc':
        productos = productos.order_by('precio')
    elif orden == 'precio_desc':
        productos = productos.order_by('-precio')
    else:
        productos = productos.order_by('-id')
    
    # Paginación
    paginator = Paginator(productos, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'COLOR_CHOICES': COLOR_CHOICES,
        'GENERO_CHOICES': GENERO_CHOICES,
        'current_filters': {
            'precio_min': precio_min,
            'precio_max': precio_max,
            'color': color,
            'genero': genero,
            'orden': orden
        },
        'template_name': 'Accesorios'
    }
    return render(request, "Accesorios.html", context)

def Contacto(request):
    if request.method == 'POST':
        form = ContactoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Tu mensaje ha sido enviado con éxito!')
            return redirect('Contacto')
    else:
        form = ContactoForm()
    
    return render(request, "Contacto.html", {'form': form})

@login_required
def Administracion(request):
    return render(request, "producto/Administracion.html")

@login_required
def marcar_procesado(request, envio_id):
    envio = get_object_or_404(EnvioModel, id=envio_id)
    envio.procesado = True
    envio.save()
    messages.success(request, f'El envío #{envio_id} ha sido marcado como procesado.')
    return redirect('FormularioEnvio')

@login_required
def FormularioEnvio(request):
    # Obtener parámetros de filtrado
    estado = request.GET.get('estado')
    marca = request.GET.get('marca')
    courier = request.GET.get('courier')
    
    # Base query
    envios = EnvioModel.objects.all().order_by('-fecha')
    
    # Aplicar filtros
    if estado:
        estado_bool = estado.lower() == 'true'
        envios = envios.filter(procesado=estado_bool)
    
    if marca:
        envios = envios.filter(marcas_compradas__icontains=marca)
    
    if courier:
        envios = envios.filter(courier=courier)
    
    # Obtener valores únicos para los filtros
    estados = [{'value': 'true', 'label': 'Procesado'}, {'value': 'false', 'label': 'No Procesado'}]
    
    # Obtener marcas únicas
    marcas = EnvioModel.objects.exclude(marcas_compradas__isnull=True)\
                               .values_list('marcas_compradas', flat=True)\
                               .distinct()
    # Convertir a lista y eliminar valores vacíos
    marcas = [m for m in marcas if m]
    
    # Obtener couriers disponibles
    couriers = EnvioModel.COURIER_CHOICES
    
    return render(request, "Usuarios/FormularioEnvio.html", {
        'envios': envios,
        'estados': estados,
        'marcas': marcas,
        'couriers': couriers,
        'current_filters': {
            'estado': estado,
            'marca': marca,
            'courier': courier
        }
    })

@login_required
def agregar_producto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            tipo = form.cleaned_data['tipo']
            genero = form.cleaned_data['genero']
            color = form.cleaned_data['color']
            cantidad = form.cleaned_data['cantidad']
            modelo = form.cleaned_data['modelo']
            precio = form.cleaned_data['precio']
            descripcion = form.cleaned_data['descripcion']
            nuevo = form.cleaned_data['nuevo']
            marca = form.cleaned_data['marca']
            imagen = form.cleaned_data['imagen']

            if tipo == 'Tronco':
                subcategoria = form.cleaned_data['subcategoria_tronco']
                Tronco.objects.create(
                    modelo=modelo,
                    precio=precio,
                    descripcion=descripcion,
                    nuevo=nuevo,
                    marca=marca,
                    imagen=imagen,
                    cantidad=cantidad,
                    genero=genero,
                    color=color,
                    subcategoria=subcategoria
                )
            elif tipo == 'Piernas':
                subcategoria = form.cleaned_data['subcategoria_piernas']
                Piernas.objects.create(
                    modelo=modelo,
                    precio=precio,
                    descripcion=descripcion,
                    nuevo=nuevo,
                    marca=marca,
                    imagen=imagen,
                    cantidad=cantidad,
                    genero=genero,
                    color=color,
                    subcategoria=subcategoria
                )
            elif tipo == 'Zapatos':
                subcategoria = form.cleaned_data['subcategoria_zapatos']
                Zapatos.objects.create(
                    modelo=modelo,
                    precio=precio,
                    descripcion=descripcion,
                    nuevo=nuevo,
                    marca=marca,
                    imagen=imagen,
                    cantidad=cantidad,
                    genero=genero,
                    color=color,
                    subcategoria=subcategoria
                )
            elif tipo == 'Complemento':
                tipo_accesorio = form.cleaned_data['tipo_accesorio']
                Complemento.objects.create(
                    modelo=modelo,
                    precio=precio,
                    descripcion=descripcion,
                    nuevo=nuevo,
                    marca=marca,
                    imagen=imagen,
                    cantidad=cantidad,
                    genero=genero,
                    color=color,
                    tipo=tipo_accesorio
                )
            
            messages.success(request, 'Producto agregado correctamente!')
            return redirect('agregar_producto')
    
    else:  # CASO GET
        form = ProductoForm()
    
    # Renderizar siempre el formulario (GET o POST inválido)
    return render(request, 'producto/agregar.html', {'form': form}) 

@login_required
def agrega_marca(request):
    if request.method == 'POST':
        form = MarcaForm(request.POST)
        if form.is_valid():
            marca = form.save(commit=False)
            
            # Asignar usuario logueado
            marca.usuario = request.user
            
            # Asignar tienda seleccionada
            tienda_id = request.POST.get('tienda')
            if tienda_id:
                try:
                    tienda = Tienda.objects.get(id=tienda_id)
                    marca.tienda = tienda
                except Tienda.DoesNotExist:
                    pass
            
            marca.save()
            messages.success(request, 'Marca agregada correctamente!')
            return redirect('agrega_marca')
    
    form = MarcaForm()
    
    # Obtener tiendas para el selector
    tiendas = Tienda.objects.all()

    # Obtener marcas por categoría
    marcas_torso = Marca.objects.filter(tronco__isnull=False).distinct()
    marcas_pantalones = Marca.objects.filter(piernas__isnull=False).distinct()
    marcas_calzado = Marca.objects.filter(zapatos__isnull=False).distinct()
    marcas_accesorios = Marca.objects.filter(complemento__isnull=False).distinct()
    
    context = {
        'form': form,
        'tiendas': tiendas, # AGREGADO: Variable solicitada
        'marcas_torso': marcas_torso,
        'marcas_pantalones': marcas_pantalones,
        'marcas_calzado': marcas_calzado,
        'marcas_accesorios': marcas_accesorios,
    }
    return render(request, "producto/agrega_marca.html", context)

def agregar_tienda(request):
    if request.method == 'POST':
        form = TiendaForm(request.POST)
        if form.is_valid():
            # Obtener datos limpios del formulario
            nombre_tienda = form.cleaned_data['nombre_tienda']
            categorias_list = form.cleaned_data['categorias'] # Esto devuelve una lista ['Torso', 'Calzado']
            
            # Convertir la lista a un string separado por comas para guardar en CharField
            categorias_str = ", ".join(categorias_list)
            
            # GUARDAR EN BASE DE DATOS
            try:
                Tienda.objects.create(
                    nombre_tienda=nombre_tienda,
                    categorias=categorias_str
                )
                messages.success(request, f'¡Tienda "{nombre_tienda}" creada con éxito!')
                return redirect('agregar_tienda')
            except Exception as e:
                messages.error(request, f'Error al guardar la tienda: {str(e)}')
    else:
        form = TiendaForm()

    return render(request, "producto/agregar_tienda.html", {'form': form})

@login_required
def modificar(request):
    productos = None
    tipo_producto = None
    marcas = Marca.objects.all()
    
    if request.method == 'POST':
        # Si es para listar productos
        if 'listar' in request.POST:
            tipo_producto = request.POST.get('tipo_producto')
            if tipo_producto == 'Tronco':
                productos = Tronco.objects.all()
            elif tipo_producto == 'Piernas':
                productos = Piernas.objects.all()
            elif tipo_producto == 'Zapatos':
                productos = Zapatos.objects.all()
            elif tipo_producto == 'Complemento':
                productos = Complemento.objects.all()
        
        # Si es para actualizar un producto
        elif 'actualizar' in request.POST:
            tipo_producto = request.POST.get('tipo_producto')
            producto_id = request.POST.get('producto_id')
            
            try:
                if tipo_producto == 'Tronco':
                    producto = Tronco.objects.get(id=producto_id)
                elif tipo_producto == 'Piernas':
                    producto = Piernas.objects.get(id=producto_id)
                elif tipo_producto == 'Zapatos':
                    producto = Zapatos.objects.get(id=producto_id)
                elif tipo_producto == 'Complemento':
                    producto = Complemento.objects.get(id=producto_id)
                
                # Actualizar campos básicos
                producto.modelo = request.POST.get('modelo', '')
                producto.precio = int(request.POST.get('precio', 0))
                producto.descripcion = request.POST.get('descripcion', '')
                producto.nuevo = 'nuevo' in request.POST
                producto.marca = Marca.objects.get(id=request.POST.get('marca'))
                producto.cantidad = int(request.POST.get('cantidad', 0))
                
                # Actualizar campos nuevos
                producto.genero = request.POST.get('genero')
                producto.color = request.POST.get('color')
                
                # Actualizar subcategoría según el tipo de producto
                if tipo_producto == 'Tronco':
                    producto.subcategoria = request.POST.get('subcategoria')
                elif tipo_producto == 'Piernas':
                    producto.subcategoria = request.POST.get('subcategoria')
                elif tipo_producto == 'Zapatos':
                    producto.subcategoria = request.POST.get('subcategoria')
                elif tipo_producto == 'Complemento':
                    producto.tipo = request.POST.get('tipo_accesorio')
                
                # Actualizar imagen si se proporcionó una nueva
                if 'imagen' in request.FILES:
                    producto.imagen = request.FILES['imagen']
                
                producto.save()
                messages.success(request, 'Producto modificado correctamente!')
                return redirect('modificar')
            
            except Exception as e:
                messages.error(request, f'Error al modificar producto: {str(e)}')
                return redirect('modificar')
    
    return render(request, "producto/modificar.html", {
        'productos': productos,
        'tipo_producto': tipo_producto,
        'marcas': marcas,
        'GENERO_CHOICES': GENERO_CHOICES,
        'COLOR_CHOICES': COLOR_CHOICES,
        'SUBCATEGORIA_TRONCO': SUBCATEGORIA_TRONCO,
        'SUBCATEGORIA_PIERNAS': SUBCATEGORIA_PIERNAS,
        'SUBCATEGORIA_ZAPATOS': SUBCATEGORIA_ZAPATOS,
    })

@login_required
def eliminar(request):
    productos = None
    tipo_producto = None
    
    if request.method == 'POST':
        # Si es para listar productos
        if 'listar' in request.POST:
            tipo_producto = request.POST.get('tipo_producto')
            if tipo_producto == 'Tronco':
                productos = Tronco.objects.all()
            elif tipo_producto == 'Piernas':
                productos = Piernas.objects.all()
            elif tipo_producto == 'Zapatos':
                productos = Zapatos.objects.all()
            elif tipo_producto == 'Complemento':
                productos = Complemento.objects.all()
        
        # Si es para eliminar productos seleccionados
        elif 'eliminar_seleccionados' in request.POST:
            tipo_producto = request.POST.get('tipo_producto')
            seleccionados = request.POST.getlist('seleccionados')
            
            if not seleccionados:
                messages.warning(request, 'No has seleccionado ningún producto para eliminar.')
            else:
                try:
                    if tipo_producto == 'Tronco':
                        Tronco.objects.filter(id__in=seleccionados).delete()
                    elif tipo_producto == 'Piernas':
                        Piernas.objects.filter(id__in=seleccionados).delete()
                    elif tipo_producto == 'Zapatos':
                        Zapatos.objects.filter(id__in=seleccionados).delete()
                    elif tipo_producto == 'Complemento':
                        Complemento.objects.filter(id__in=seleccionados).delete()
                    
                    messages.success(request, f'Se han eliminado {len(seleccionados)} producto(s) correctamente.')
                    return redirect('eliminar')
                except Exception as e:
                    messages.error(request, f'Error al eliminar productos: {str(e)}')
                    return redirect('eliminar')
        
        # Si es para eliminar todos los productos del tipo seleccionado
        elif 'eliminar_todos' in request.POST:
            tipo_producto = request.POST.get('tipo_producto')
            try:
                if tipo_producto == 'Tronco':
                    count = Tronco.objects.count()
                    Tronco.objects.all().delete()
                elif tipo_producto == 'Piernas':
                    count = Piernas.objects.count()
                    Piernas.objects.all().delete()
                elif tipo_producto == 'Zapatos':
                    count = Zapatos.objects.count()
                    Zapatos.objects.all().delete()
                elif tipo_producto == 'Complemento':
                    count = Complemento.objects.count()
                    Complemento.objects.all().delete()
                
                messages.success(request, f'Se han eliminado {count} producto(s) correctamente.')
                return redirect('eliminar')
            except Exception as e:
                messages.error(request, f'Error al eliminar todos los productos: {str(e)}')
                return redirect('eliminar')
    
    return render(request, "producto/eliminar.html", {
        'productos': productos,
        'tipo_producto': tipo_producto,
    })

@login_required
def Admin_User(request):
    users = User.objects.all().order_by('id')
    
    if request.method == 'POST':
        if 'delete_user' in request.POST:
            user_id = request.POST.get('user_id')
            try:
                user = User.objects.get(id=user_id)
                user.delete()
                messages.success(request, f'Usuario {user.username} eliminado correctamente.')
            except User.DoesNotExist:
                messages.error(request, 'El usuario no existe.')
            return redirect('Admin_User')
            
        elif 'update_user' in request.POST:
            user_id = request.POST.get('user_id')
            username = request.POST.get(f'username_{user_id}')
            email = request.POST.get(f'email_{user_id}')
            first_name = request.POST.get(f'first_name_{user_id}')
            last_name = request.POST.get(f'last_name_{user_id}')
            is_staff = request.POST.get(f'is_staff_{user_id}') == 'on'
            
            try:
                user = User.objects.get(id=user_id)
                user.username = username
                user.email = email
                user.first_name = first_name
                user.last_name = last_name
                user.is_staff = is_staff
                user.save()
                messages.success(request, f'Usuario {user.username} actualizado correctamente.')
            except User.DoesNotExist:
                messages.error(request, 'El usuario no existe.')
            return redirect('Admin_User')
    
    return render(request, "Usuarios/Admin_User.html", {'users': users})

def IniciarSesion(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            return redirect('index')  # Redirige a la página principal después del login
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')
    
    return render(request, "Usuarios/IniciarSesion.html")

def cerrar_sesion(request):
    logout(request)
    return redirect('index')

def Registrarse(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.first_name = request.POST.get('first_name')
            user.last_name = request.POST.get('last_name')
            user.email = request.POST.get('email')
            user.save()

            recibir_notificaciones = request.POST.get('recibir_notificaciones') == 'on'
            PreferenciasUsuario.objects.create(user=user, recibir_notificaciones=recibir_notificaciones)

            messages.success(request, f'¡Cuenta creada para {user.username}!')
            return redirect('IniciarSesion')
    else:
        form = UserCreationForm()

    # 🔹 Aplica clases de Bootstrap a todos los campos
    for field in form.fields.values():
        field.widget.attrs['class'] = 'form-control'

    return render(request, "Usuarios/Registrarse.html", {'form': form})


def QuienesSomos(request):
    return render(request, "Footer/QuienesSomos.html")

def SeSocio(request):
    return render(request, "Footer/SeSocio.html")

def TrabajaConNosotros(request):
    if request.method == 'POST':
        form = TrabajaConNosotrosForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Tu solicitud ha sido enviada con éxito! Nos pondremos en contacto contigo pronto.')
            return redirect('TrabajaConNosotros')
    else:
        form = TrabajaConNosotrosForm()
    
    return render(request, "Footer/TrabajaConNosotros.html", {'form': form})

def TerminosCondiciones(request):
    return render(request, "Footer/TerminosCondiciones.html")

def SeSocio(request):
    if request.method == 'POST':
        form = SeSocioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Tu solicitud ha sido enviada con éxito! Nos pondremos en contacto contigo pronto.')
            return redirect('SeSocio')
    else:
        form = SeSocioForm()
    
    return render(request, "Footer/SeSocio.html", {'form': form})

def commit_pay(request):
    return render(request, "Ventas/commit_pay.html")

def pago_error(request):
    return render(request, "Ventas/pago_error.html")

def send_pay(request):
    return render(request, "Ventas/send_pay.html")

def exportar_envios_no_procesados(request):
    # Obtener todos los envíos no procesados
    envios = EnvioModel.objects.filter(procesado=False).order_by('fecha')
    
    # Crear un libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Envios no procesados"
    
    # Encabezados
    headers = [
        'ID', 'Nombre', 'Email', 'Teléfono', 'Dirección', 
        'Ciudad', 'Región', 'Código Postal', 'Courier', 
        'Subtotal', 'Costo Envío', 'Total', 'Marcas', 'Fecha'
    ]
    
    # Estilos para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment = Alignment(horizontal='center', vertical='center')
    
    # Escribir encabezados
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}1"]
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        ws.column_dimensions[col_letter].width = 20
    
    # Escribir datos
    for row_num, envio in enumerate(envios, 2):
        ws.cell(row=row_num, column=1, value=envio.id)
        ws.cell(row=row_num, column=2, value=envio.nombre)
        ws.cell(row=row_num, column=3, value=envio.email)
        ws.cell(row=row_num, column=4, value=envio.telefono)
        ws.cell(row=row_num, column=5, value=envio.direccion)
        ws.cell(row=row_num, column=6, value=envio.ciudad)
        ws.cell(row=row_num, column=7, value=envio.get_region_display())
        ws.cell(row=row_num, column=8, value=envio.codigo_postal)
        ws.cell(row=row_num, column=9, value=envio.get_courier_display())
        ws.cell(row=row_num, column=10, value=envio.monto_carrito)
        ws.cell(row=row_num, column=11, value=envio.costo_envio)
        ws.cell(row=row_num, column=12, value=envio.total_pagar)
        ws.cell(row=row_num, column=13, value=envio.marcas_compradas)
        ws.cell(row=row_num, column=14, value=envio.fecha.strftime("%d/%m/%Y %H:%M"))
    
    # Preparar la respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="envios_no_procesados.xlsx"'
    
    # Guardar el libro en la respuesta
    wb.save(response)
    
    return response

