from django.shortcuts import render, redirect, get_object_or_404
from django.http import HttpResponse
from .forms import MarcaForm, ProductoForm, ContactoForm, TrabajaConNosotrosForm, SeSocioForm, EnvioForm, TiendaForm
from .models import (
    GENERO_CHOICES, 
    COLOR_CHOICES, 
    SUBCATEGORIA_TRONCO, 
    SUBCATEGORIA_PIERNAS, 
    SUBCATEGORIA_ZAPATOS,
    Marca, Tronco, Piernas, Zapatos, Complemento, HistorialEliminacion, HistorialModificacion, Envio as EnvioModel, Tienda
)
from .models import Tienda, PreferenciasUsuario, User
from django.contrib import messages
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth.forms import UserCreationForm, UserChangeForm
from django.http import JsonResponse
from django.views.decorators.http import require_POST
from django.contrib.auth.decorators import login_required
import json
from django.contrib.auth.decorators import user_passes_test
from django.conf import settings
from .transbank_utils import transbank_create, transbank_commit, transbank_reverse_or_cancel
import datetime as dt
from django.views.decorators.csrf import csrf_exempt
from django.core.paginator import Paginator
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

# Create your views here.
TEMPLATE_DIRS = (
    'os.path.join(BASE_DIR, "templates")'
)

def index(request):
    # Obtener el producto más caro de cada categoría
    tronco_mas_caro = Tronco.objects.order_by('-precio').first()
    piernas_mas_caro = Piernas.objects.order_by('-precio').first()
    zapatos_mas_caro = Zapatos.objects.order_by('-precio').first()
    complemento_mas_caro = Complemento.objects.order_by('-precio').first()
    
    return render(request, "index.html", {
        'tronco_mas_caro': tronco_mas_caro,
        'piernas_mas_caro': piernas_mas_caro,
        'zapatos_mas_caro': zapatos_mas_caro,
        'complemento_mas_caro': complemento_mas_caro
    })

def envio_view(request):
    print("Accediendo a envio_view")  # Depuración
    cart = request.session.get('cart', {})
    print(f"Carrito en sesión: {cart}")  # Depuración
    cart = request.session.get('cart', {})
    if not cart:
        messages.error(request, 'Tu carrito está vacío. Agrega productos antes de proceder al envío.')
        return redirect('carrito')
    total_carrito = request.session.get('total_carrito', 0)
    
    marcas = set()
    for tipo, items in cart.items():
        for id_producto in items.keys():
            try:
                modelo = None
                if tipo == 'Tronco':
                    modelo = Tronco.objects.get(id=id_producto)
                elif tipo == 'Piernas':
                    modelo = Piernas.objects.get(id=id_producto)
                elif tipo == 'Zapatos':
                    modelo = Zapatos.objects.get(id=id_producto)
                elif tipo == 'Complemento':
                    modelo = Complemento.objects.get(id=id_producto)
                
                if modelo and modelo.marca:
                    marcas.add(modelo.marca.nombre)
            except:
                continue
    
    marcas_str = ", ".join(marcas) if marcas else ""

    if request.method == 'POST':
        form = EnvioForm(request.POST)
        if form.is_valid():
            envio = form.save(commit=False)
            envio.monto_carrito = total_carrito
            envio.costo_envio = envio.calcular_costo_envio()
            envio.total_pagar = envio.monto_carrito + envio.costo_envio
            envio.marcas_compradas = marcas_str  # Asignar las marcas
            
            if request.user.is_authenticated:
                envio.usuario = request.user
                
            envio.save()
            request.session['envio_id'] = envio.id
            return redirect('Pago')
        else:
            messages.error(request, 'Por Favor, corrige los errores en el formulario')
    else:
        form = EnvioForm()
    
    return render(request, "Ventas/Envio.html", {
        'form': form,
        'total_carrito': total_carrito
    })

def pago_view(request):
    envio_id = request.session.get('envio_id')
    if not envio_id:
        return redirect('carrito')
    
    try:
        envio = EnvioModel.objects.get(id=envio_id)
    except EnvioModel.DoesNotExist:
        return redirect('carrito')
    
    if request.method == 'POST':
        # Procesar el pago y descontar del stock
        cart = request.session.get('cart', {})
        
        for product_type, products in cart.items():
            model = None
            if product_type == 'Tronco':
                model = Tronco
            elif product_type == 'Piernas':
                model = Piernas
            elif product_type == 'Zapatos':
                model = Zapatos
            elif product_type == 'Complemento':
                model = Complemento
            
            if model:
                for product_id, quantity in products.items():
                    try:
                        product = model.objects.get(id=product_id)
                        if product.cantidad >= quantity:
                            product.cantidad -= quantity
                            product.save()
                        else:
                            messages.error(request, f'No hay suficiente stock para {product.modelo}')
                            return redirect('carrito')
                    except model.DoesNotExist:
                        continue
            # Crear transacción en Transbank
            buy_order = f"USTORE_{envio_id}"
            session_id = request.session.session_key
            amount = envio.total_pagar
            return_url = settings.TRANSBANK_RETURN_URL

            body = {
                "buy_order": buy_order,
                "session_id": session_id,
                "amount": amount,
                "return_url": return_url
            }

            response = transbank_create(body)
            if response.status_code == 200:
                transbank_data = response.json()
                return render(request, 'Ventas/send_pay.html', {
                    'transbank': transbank_data,
                    'amount': amount
                })
            else:
                return render(request, 'Ventas/pago_error.html', {
                    'error': 'Error al crear la transacción en Transbank'
                })
        
        # Limpiar el carrito después del pago
        if 'cart' in request.session:
            del request.session['cart']
        if 'total_carrito' in request.session:
            del request.session['total_carrito']
        if 'envio_id' in request.session:
            del request.session['envio_id']
        
        messages.success(request, '¡Pago realizado con éxito! Tu pedido ha sido procesado.')
        return redirect('index')
    
    return render(request, "Ventas/Pago.html", {
        'envio': envio
    })

@csrf_exempt
def commit_pay_view(request):
    transaction_detail = None
    try:
        tokenws = request.GET.get('token_ws') or request.POST.get('token_ws')
        
        if tokenws:
            response = transbank_commit(tokenws)
            
            if response.status_code == 200:
                response_data = response.json()
                status = response_data.get('status')
                response_code = response_data.get('response_code')
                
                if status == 'AUTHORIZED' and response_code == 0:
                    # Pago exitoso: descontar stock
                    cart = request.session.get('cart', {})
                    for product_type, products in cart.items():
                        model = None
                        if product_type == 'Tronco':
                            model = Tronco
                        elif product_type == 'Piernas':
                            model = Piernas
                        elif product_type == 'Zapatos':
                            model = Zapatos
                        elif product_type == 'Complemento':
                            model = Complemento
                        
                        if model:
                            for product_id, quantity in products.items():
                                try:
                                    product = model.objects.get(id=product_id)
                                    if product.cantidad >= quantity:
                                        product.cantidad -= quantity
                                        product.save()
                                    else:
                                        # Reversar pago si no hay stock suficiente
                                        amount = response_data.get('amount')
                                        reverse_response = transbank_reverse_or_cancel(tokenws, amount)
                                        return render(request, 'Ventas/pago_error.html', {
                                            'error': f'No hay suficiente stock para {product.modelo}'
                                        })
                                except model.DoesNotExist:
                                    pass
                    
                    # Limpiar carrito después del pago exitoso
                    if 'cart' in request.session:
                        del request.session['cart']
                    if 'total_carrito' in request.session:
                        del request.session['total_carrito']
                    if 'envio_id' in request.session:
                        del request.session['envio_id']
                    
                    # Preparar datos para mostrar
                    state = 'ACEPTADO'
                    pay_type = 'Tarjeta de Crédito' if response_data.get('payment_type_code') == 'VC' else 'Tarjeta de Débito'
                    amount = int(response_data.get('amount', 0))
                    amount_formatted = f'{amount:,.0f}'.replace(',', '.')
                    transaction_date = dt.datetime.strptime(response_data['transaction_date'], '%Y-%m-%dT%H:%M:%S.%fZ')
                    transaction_date = transaction_date.strftime('%d-%m-%Y %H:%M:%S')
                    
                    transaction_detail = {
                        'card_number': response_data['card_detail']['card_number'],
                        'transaction_date': transaction_date,
                        'state': state,
                        'pay_type': pay_type,
                        'amount': amount_formatted,
                        'authorization_code': response_data['authorization_code'],
                        'buy_order': response_data['buy_order'],
                    }
                else:
                    # Pago rechazado
                    state = 'RECHAZADO'
                    pay_type = 'Tarjeta de Crédito' if response_data.get('payment_type_code') == 'VC' else 'Tarjeta de Débito'
                    amount = int(response_data.get('amount', 0))
                    amount_formatted = f'{amount:,.0f}'.replace(',', '.')
                    transaction_date = dt.datetime.strptime(response_data['transaction_date'], '%Y-%m-%dT%H:%M:%S.%fZ')
                    transaction_date = transaction_date.strftime('%d-%m-%Y %H:%M:%S')
                    
                    # Intentar reversar el pago
                    reverse_response = transbank_reverse_or_cancel(tokenws, amount)
                    
                    transaction_detail = {
                        'card_number': response_data['card_detail']['card_number'],
                        'transaction_date': transaction_date,
                        'state': state,
                        'pay_type': pay_type,
                        'amount': amount_formatted,
                        'authorization_code': response_data['authorization_code'],
                        'buy_order': response_data['buy_order'],
                    }
        
        return render(request, 'Ventas/commit_pay.html', {
            'transaction_detail': transaction_detail
        })
    
    except Exception as e:
        return render(request, 'Ventas/pago_error.html', {
            'error': f'Error en el proceso de pago: {str(e)}'
        })

@login_required
def carrito(request):
    cart = request.session.get('cart', {})
    items = []
    total = 0
    
    for product_type, products in cart.items():
        model = None
        if product_type == 'Tronco':
            model = Tronco
        elif product_type == 'Piernas':
            model = Piernas
        elif product_type == 'Zapatos':
            model = Zapatos
        elif product_type == 'Complemento':
            model = Complemento
        
        if model:
            for product_id, quantity in products.items():
                try:
                    product = model.objects.get(id=product_id)
                    
                    # MOVER LA VERIFICACIÓN DE STOCK DENTRO DEL BLOQUE
                    if quantity > product.cantidad:
                        messages.warning(
                            request, 
                            f'No hay suficiente stock de {product.modelo}. Disponible: {product.cantidad}'
                        )
                    
                    subtotal = product.precio * quantity
                    total += subtotal
                    items.append({
                        'producto': product,
                        'tipo': product_type,
                        'cantidad': quantity,
                        'subtotal': subtotal
                    })
                except model.DoesNotExist:
                    continue
                    
    # Guardar total en sesión para usar en envío
    request.session['total_carrito'] = total

    return render(request, 'Ventas/carrito.html', {
        'items': items,
        'total': total
    })

def get_cart_data(request):
    """
    Vista API para devolver los datos del carrito en JSON
    para el sidebar lateral.
    """
    cart = request.session.get('cart', {})
    items_data = []
    total = 0
    
    for product_type, products in cart.items():
        model = None
        if product_type == 'Tronco':
            model = Tronco
        elif product_type == 'Piernas':
            model = Piernas
        elif product_type == 'Zapatos':
            model = Zapatos
        elif product_type == 'Complemento':
            model = Complemento
        
        if model:
            for product_id, quantity in products.items():
                try:
                    product = model.objects.get(id=product_id)
                    subtotal = product.precio * quantity
                    total += subtotal
                    
                    # Preparar imagen
                    img_url = ""
                    if product.imagen:
                        img_url = product.imagen.url
                    
                    items_data.append({
                        'id': product.id,
                        'modelo': product.modelo,
                        'precio': product.precio,
                        'cantidad': quantity,
                        'subtotal': subtotal,
                        'imagen': img_url,
                        'tipo': product_type,
                        'marca': product.marca.nombre if product.marca else ""
                    })
                except model.DoesNotExist:
                    continue
    
    return JsonResponse({
        'success': True,
        'items': items_data,
        'total': total,
        'total_items': len(items_data)
    })

@require_POST
@login_required
def update_cart(request):
    try:
        data = json.loads(request.body)
        product_id = str(data.get('product_id'))
        product_type = data.get('product_type')
        action = data.get('action')
        
        # Obtener el modelo del producto
        model = None
        if product_type == 'Tronco':
            model = Tronco
        elif product_type == 'Piernas':
            model = Piernas
        elif product_type == 'Zapatos':
            model = Zapatos
        elif product_type == 'Complemento':
            model = Complemento
        
        if not model:
            return JsonResponse({'success': False, 'message': 'Tipo de producto no válido'})
        
        # Obtener el producto
        try:
            product = model.objects.get(id=product_id)
        except model.DoesNotExist:
            return JsonResponse({'success': False, 'message': 'Producto no encontrado'})
        
        if not request.session.get('cart'):
            request.session['cart'] = {}
        
        cart = request.session['cart']
        
        if action == 'add':
            # Verificar stock disponible
            current_quantity = cart.get(product_type, {}).get(product_id, 0)
            if current_quantity + 1 > product.cantidad:
                return JsonResponse({
                    'success': False, 
                    'message': f'No hay suficiente stock. Disponible: {product.cantidad}'
                })
            
            if product_type not in cart:
                cart[product_type] = {}
            
            cart[product_type][product_id] = current_quantity + 1
            request.session.modified = True
            
        elif action == 'update':
            quantity = int(data.get('quantity', 1))
            if product_type in cart and product_id in cart[product_type]:
                cart[product_type][product_id] = quantity
                request.session.modified = True
            else:
                return JsonResponse({'success': False, 'message': 'Producto no encontrado en el carrito'})
        
        elif action == 'remove':
            if product_type in cart and product_id in cart[product_type]:
                del cart[product_type][product_id]
                # Eliminar la categoría si queda vacía
                if not cart[product_type]:
                    del cart[product_type]
                request.session.modified = True
            else:
                return JsonResponse({'success': False, 'message': 'Producto no encontrado en el carrito'})
        else:
            return JsonResponse({'success': False, 'message': 'Acción no válida'})
        
        # Calcular el total de items en el carrito
        total_items = sum(sum(products.values()) for products in cart.values())
        
        return JsonResponse({
            'success': True,
            'cart_count': total_items
        })
    
    except Exception as e:
        return JsonResponse({'success': False, 'message': str(e)})

def Torso(request):
    # Obtener parámetros de filtrado
    precio_min = request.GET.get('precio_min')
    precio_max = request.GET.get('precio_max')
    color = request.GET.get('color')
    genero = request.GET.get('genero')
    orden = request.GET.get('orden', 'reciente')  # Por defecto: ordenar por más reciente
    
    # Base query
    productos = Tronco.objects.all()
    
    # Aplicar filtros
    if precio_min:
        productos = productos.filter(precio__gte=precio_min)
    if precio_max:
        productos = productos.filter(precio__lte=precio_max)
    if color:
        productos = productos.filter(color=color)
    if genero:
        productos = productos.filter(genero=genero)
    
    # Aplicar ordenamiento
    if orden == 'precio_asc':
        productos = productos.order_by('precio')
    elif orden == 'precio_desc':
        productos = productos.order_by('-precio')
    else:  # 'reciente' o cualquier otro
        productos = productos.order_by('-id')  # Orden por defecto: más reciente primero
    
    # Paginación
    paginator = Paginator(productos, 12)  # 12 productos por página
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'COLOR_CHOICES': COLOR_CHOICES,
        'GENERO_CHOICES': GENERO_CHOICES,
        'current_filters': {
            'precio_min': precio_min,
            'precio_max': precio_max,
            'color': color,
            'genero': genero,
            'orden': orden
        },
        'template_name': 'Torso'  # Nueva variable
    }
    return render(request, "Torso.html", context)

def Pantalones(request):
    # Obtener parámetros de filtrado
    precio_min = request.GET.get('precio_min')
    precio_max = request.GET.get('precio_max')
    color = request.GET.get('color')
    genero = request.GET.get('genero')
    orden = request.GET.get('orden', 'reciente')
    
    # Base query
    productos = Piernas.objects.all()
    
    # Aplicar filtros
    if precio_min:
        productos = productos.filter(precio__gte=precio_min)
    if precio_max:
        productos = productos.filter(precio__lte=precio_max)
    if color:
        productos = productos.filter(color=color)
    if genero:
        productos = productos.filter(genero=genero)
    
    # Aplicar ordenamiento
    if orden == 'precio_asc':
        productos = productos.order_by('precio')
    elif orden == 'precio_desc':
        productos = productos.order_by('-precio')
    else:
        productos = productos.order_by('-id')
    
    # Paginación
    paginator = Paginator(productos, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'COLOR_CHOICES': COLOR_CHOICES,
        'GENERO_CHOICES': GENERO_CHOICES,
        'current_filters': {
            'precio_min': precio_min,
            'precio_max': precio_max,
            'color': color,
            'genero': genero,
            'orden': orden
        },
        'template_name': 'Pantalones' 
    }
    return render(request, "Pantalones.html", context)

def Calzado(request):
    # Obtener parámetros de filtrado
    precio_min = request.GET.get('precio_min')
    precio_max = request.GET.get('precio_max')
    color = request.GET.get('color')
    genero = request.GET.get('genero')
    orden = request.GET.get('orden', 'reciente')
    
    # Base query
    productos = Zapatos.objects.all()
    
    # Aplicar filtros
    if precio_min:
        productos = productos.filter(precio__gte=precio_min)
    if precio_max:
        productos = productos.filter(precio__lte=precio_max)
    if color:
        productos = productos.filter(color=color)
    if genero:
        productos = productos.filter(genero=genero)
    
    # Aplicar ordenamiento
    if orden == 'precio_asc':
        productos = productos.order_by('precio')
    elif orden == 'precio_desc':
        productos = productos.order_by('-precio')
    else:
        productos = productos.order_by('-id')
    
    # Paginación
    paginator = Paginator(productos, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'COLOR_CHOICES': COLOR_CHOICES,
        'GENERO_CHOICES': GENERO_CHOICES,
        'current_filters': {
            'precio_min': precio_min,
            'precio_max': precio_max,
            'color': color,
            'genero': genero,
            'orden': orden
        },
        'template_name': 'Calzado'
    }
    return render(request, "Calzado.html", context)

def Accesorios(request):
    # Obtener parámetros de filtrado
    precio_min = request.GET.get('precio_min')
    precio_max = request.GET.get('precio_max')
    color = request.GET.get('color')
    genero = request.GET.get('genero')
    orden = request.GET.get('orden', 'reciente')
    
    # Base query
    productos = Complemento.objects.all()
    
    # Aplicar filtros
    if precio_min:
        productos = productos.filter(precio__gte=precio_min)
    if precio_max:
        productos = productos.filter(precio__lte=precio_max)
    if color:
        productos = productos.filter(color=color)
    if genero:
        productos = productos.filter(genero=genero)
    
    # Aplicar ordenamiento
    if orden == 'precio_asc':
        productos = productos.order_by('precio')
    elif orden == 'precio_desc':
        productos = productos.order_by('-precio')
    else:
        productos = productos.order_by('-id')
    
    # Paginación
    paginator = Paginator(productos, 12)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'COLOR_CHOICES': COLOR_CHOICES,
        'GENERO_CHOICES': GENERO_CHOICES,
        'current_filters': {
            'precio_min': precio_min,
            'precio_max': precio_max,
            'color': color,
            'genero': genero,
            'orden': orden
        },
        'template_name': 'Accesorios'
    }
    return render(request, "Accesorios.html", context)

def Contacto(request):
    if request.method == 'POST':
        form = ContactoForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Tu mensaje ha sido enviado con éxito!')
            return redirect('Contacto')
    else:
        form = ContactoForm()
    
    return render(request, "Contacto.html", {'form': form})

@login_required
def Administracion(request):
    return render(request, "producto/Administracion.html")

@login_required
def marcar_procesado(request, envio_id):
    envio = get_object_or_404(EnvioModel, id=envio_id)
    envio.procesado = True
    envio.save()
    messages.success(request, f'El envío #{envio_id} ha sido marcado como procesado.')
    return redirect('FormularioEnvio')

@login_required
def FormularioEnvio(request):
    # Obtener parámetros de filtrado
    estado = request.GET.get('estado')
    marca = request.GET.get('marca')
    courier = request.GET.get('courier')
    
    # Base query
    envios = EnvioModel.objects.all().order_by('-fecha')
    
    # Aplicar filtros
    if estado:
        estado_bool = estado.lower() == 'true'
        envios = envios.filter(procesado=estado_bool)
    
    if marca:
        envios = envios.filter(marcas_compradas__icontains=marca)
    
    if courier:
        envios = envios.filter(courier=courier)
    
    # Obtener valores únicos para los filtros
    estados = [{'value': 'true', 'label': 'Procesado'}, {'value': 'false', 'label': 'No Procesado'}]
    
    # Obtener marcas únicas
    marcas = EnvioModel.objects.exclude(marcas_compradas__isnull=True)\
                               .values_list('marcas_compradas', flat=True)\
                               .distinct()
    # Convertir a lista y eliminar valores vacíos
    marcas = [m for m in marcas if m]
    
    # Obtener couriers disponibles
    couriers = EnvioModel.COURIER_CHOICES
    
    return render(request, "Usuarios/FormularioEnvio.html", {
        'envios': envios,
        'estados': estados,
        'marcas': marcas,
        'couriers': couriers,
        'current_filters': {
            'estado': estado,
            'marca': marca,
            'courier': courier
        }
    })

@login_required
def agregar_producto(request):
    if request.method == 'POST':
        form = ProductoForm(request.POST, request.FILES)
        if form.is_valid():
            # Extracción de datos comunes
            tipo = form.cleaned_data['tipo']
            modelo = form.cleaned_data['modelo']
            precio = form.cleaned_data['precio']
            descripcion = form.cleaned_data['descripcion']
            nuevo = form.cleaned_data['nuevo']
            marca = form.cleaned_data['marca']
            imagen = form.cleaned_data['imagen']
            cantidad = form.cleaned_data['cantidad']
            genero = form.cleaned_data['genero']
            color = form.cleaned_data['color']
            
            # Captura del usuario que realiza la carga
            usuario_actual = request.user

            if tipo == 'Tronco':
                subcategoria = form.cleaned_data['subcategoria_tronco']
                Tronco.objects.create(
                    modelo=modelo,
                    precio=precio,
                    descripcion=descripcion,
                    nuevo=nuevo,
                    marca=marca,
                    imagen=imagen,
                    cantidad=cantidad,
                    genero=genero,
                    color=color,
                    subcategoria=subcategoria,
                    usuario=usuario_actual
                )
            elif tipo == 'Piernas':
                subcategoria = form.cleaned_data['subcategoria_piernas']
                Piernas.objects.create(
                    modelo=modelo,
                    precio=precio,
                    descripcion=descripcion,
                    nuevo=nuevo,
                    marca=marca,
                    imagen=imagen,
                    cantidad=cantidad,
                    genero=genero,
                    color=color,
                    subcategoria=subcategoria,
                    usuario=usuario_actual
                )
            elif tipo == 'Zapatos':
                subcategoria = form.cleaned_data['subcategoria_zapatos']
                Zapatos.objects.create(
                    modelo=modelo,
                    precio=precio,
                    descripcion=descripcion,
                    nuevo=nuevo,
                    marca=marca,
                    imagen=imagen,
                    cantidad=cantidad,
                    genero=genero,
                    color=color,
                    subcategoria=subcategoria,
                    usuario=usuario_actual
                )
            elif tipo == 'Complemento':
                subcategoria = form.cleaned_data['tipo_accesorio']
                Complemento.objects.create(
                    modelo=modelo,
                    precio=precio,
                    descripcion=descripcion,
                    nuevo=nuevo,
                    marca=marca,
                    imagen=imagen,
                    cantidad=cantidad,
                    genero=genero,
                    color=color,
                    subcategoria=subcategoria,
                    usuario=usuario_actual
                )
            
            return redirect('Administracion')
    else:
        form = ProductoForm()

    return render(request, 'producto/agregar.html', {'form': form})

@login_required
def agrega_marca(request):
    if request.method == 'POST':
        form = MarcaForm(request.POST)
        if form.is_valid():
            marca = form.save(commit=False)
            
            # 1. Asignar Usuario
            marca.usuario = request.user
            
            # 2. Guardamos también la tienda histórica (opcional, pero útil como respaldo)
            if hasattr(request.user, 'preferencias') and request.user.preferencias.tienda_asociada:
                marca.tienda = request.user.preferencias.tienda_asociada
            
            marca.save()
            messages.success(request, 'Marca registrada correctamente.')
            return redirect('agrega_marca')
    else:
        form = MarcaForm()
    
    # CONSULTA ACTUALIZADA:
    # Traemos 'usuario__preferencias__tienda_asociada' para ver la tienda EN VIVO del usuario.
    # Mantenemos 'tienda' como respaldo por si el usuario fue eliminado.
    marcas = Marca.objects.all().select_related(
        'usuario__preferencias__tienda_asociada', 
        'tienda'
    ).order_by('-id')

    data = {
        'form': form,
        'marcas': marcas,
    }

    return render(request, "producto/agrega_marca.html", data)

def agregar_tienda(request):
    if request.method == 'POST':
        form = TiendaForm(request.POST)
        if form.is_valid():
            # Obtener datos limpios del formulario
            nombre_tienda = form.cleaned_data['nombre_tienda']
            categorias_list = form.cleaned_data['categorias'] # Esto devuelve una lista ['Torso', 'Calzado']
            
            # Convertir la lista a un string separado por comas para guardar en CharField
            categorias_str = ", ".join(categorias_list)
            
            # GUARDAR EN BASE DE DATOS
            try:
                Tienda.objects.create(
                    nombre_tienda=nombre_tienda,
                    categorias=categorias_str
                )
                messages.success(request, f'¡Tienda "{nombre_tienda}" creada con éxito!')
                return redirect('agregar_tienda')
            except Exception as e:
                messages.error(request, f'Error al guardar la tienda: {str(e)}')
    else:
        form = TiendaForm()

    return render(request, "producto/agregar_tienda.html", {'form': form})

@login_required
def modificar(request):
    productos = None
    tipo_producto = None
    marcas = Marca.objects.all()
    # Obtener los últimos 20 registros del historial para la tabla
    historial = HistorialModificacion.objects.all().order_by('-fecha')[:20]
    
    if request.method == 'POST':
        # Si es para listar productos
        if 'listar' in request.POST:
            tipo_producto = request.POST.get('tipo_producto')
            if tipo_producto == 'Tronco':
                productos = Tronco.objects.all()
            elif tipo_producto == 'Piernas':
                productos = Piernas.objects.all()
            elif tipo_producto == 'Zapatos':
                productos = Zapatos.objects.all()
            elif tipo_producto == 'Complemento':
                productos = Complemento.objects.all()
        
        # Si es para actualizar un producto
        elif 'actualizar' in request.POST:
            tipo_producto = request.POST.get('tipo_producto')
            producto_id = request.POST.get('producto_id')
            
            try:
                # Selección del modelo según el tipo
                if tipo_producto == 'Tronco':
                    producto = Tronco.objects.get(id=producto_id)
                elif tipo_producto == 'Piernas':
                    producto = Piernas.objects.get(id=producto_id)
                elif tipo_producto == 'Zapatos':
                    producto = Zapatos.objects.get(id=producto_id)
                elif tipo_producto == 'Complemento':
                    producto = Complemento.objects.get(id=producto_id)
                
                # Actualizar campos básicos
                producto.modelo = request.POST.get('modelo', '')
                producto.precio = int(request.POST.get('precio', 0))
                producto.descripcion = request.POST.get('descripcion', '')
                producto.nuevo = 'nuevo' in request.POST
                producto.marca = Marca.objects.get(id=request.POST.get('marca'))
                producto.cantidad = int(request.POST.get('cantidad', 0))
                
                # Actualizar campos de características
                producto.genero = request.POST.get('genero')
                producto.color = request.POST.get('color')
                producto.usuario = request.user
                
                # Actualizar subcategoría según el tipo de producto
                if tipo_producto == 'Tronco':
                    producto.subcategoria = request.POST.get('subcategoria')
                elif tipo_producto == 'Piernas':
                    producto.subcategoria = request.POST.get('subcategoria')
                elif tipo_producto == 'Zapatos':
                    producto.subcategoria = request.POST.get('subcategoria')
                elif tipo_producto == 'Complemento':
                    producto.subcategoria = request.POST.get('tipo_accesorio')
                
                # Actualizar imagen
                if 'imagen' in request.FILES:
                    producto.imagen = request.FILES['imagen']
                
                producto.save()

                # --- REGISTRO EN HISTORIAL ---
                # Determinamos el nombre de la tienda asociada a la marca del producto
                nombre_tienda = "Sin Tienda"
                if producto.marca and producto.marca.tienda:
                    nombre_tienda = producto.marca.tienda.nombre_tienda
                
                HistorialModificacion.objects.create(
                    usuario=request.user,
                    prenda=producto.modelo,
                    tienda=nombre_tienda
                )
                # -----------------------------
                
                messages.success(request, f'Producto modificado exitosamente por: {request.user.username.upper()}')
                return redirect('modificar')
            
            except Exception as e:
                messages.error(request, f'Error al modificar producto: {str(e)}')
                return redirect('modificar')
    
    return render(request, "producto/modificar.html", {
        'productos': productos,
        'tipo_producto': tipo_producto,
        'marcas': marcas,
        'GENERO_CHOICES': GENERO_CHOICES,
        'COLOR_CHOICES': COLOR_CHOICES,
        'SUBCATEGORIA_TRONCO': SUBCATEGORIA_TRONCO,
        'SUBCATEGORIA_PIERNAS': SUBCATEGORIA_PIERNAS,
        'SUBCATEGORIA_ZAPATOS': SUBCATEGORIA_ZAPATOS,
        'historial': historial, # Enviamos el historial al template
    })

@login_required
def eliminar(request):
    productos = None
    tipo_producto = None
    
    # Consulta crítica: Carga el historial para enviarlo al HTML
    historial_bajas = HistorialEliminacion.objects.all().order_by('-fecha')[:50]
    
    if request.method == 'POST':
        # --- CASO 1: LISTAR PRODUCTOS ---
        if 'listar' in request.POST:
            tipo_producto = request.POST.get('tipo_producto')
            if tipo_producto == 'Tronco':
                productos = Tronco.objects.all()
            elif tipo_producto == 'Piernas':
                productos = Piernas.objects.all()
            elif tipo_producto == 'Zapatos':
                productos = Zapatos.objects.all()
            elif tipo_producto == 'Complemento':
                productos = Complemento.objects.all()
        
        # --- CASO 2: ELIMINAR SELECCIONADOS ---
        elif 'eliminar_seleccionados' in request.POST:
            tipo_producto = request.POST.get('tipo_producto')
            seleccionados = request.POST.getlist('seleccionados')
            
            if not seleccionados:
                messages.warning(request, 'No has seleccionado ningún producto para eliminar.')
            else:
                try:
                    model_class = None
                    if tipo_producto == 'Tronco': model_class = Tronco
                    elif tipo_producto == 'Piernas': model_class = Piernas
                    elif tipo_producto == 'Zapatos': model_class = Zapatos
                    elif tipo_producto == 'Complemento': model_class = Complemento
                    
                    if model_class:
                        # Recuperar objetos antes de borrar para guardarlos en historial
                        items_a_borrar = model_class.objects.filter(id__in=seleccionados)
                        
                        historial_objs = []
                        for item in items_a_borrar:
                            marca_nombre = item.marca.nombre if item.marca else "Sin Marca"
                            historial_objs.append(HistorialEliminacion(
                                usuario=request.user,
                                producto_modelo=item.modelo,
                                marca=marca_nombre,
                                tipo_producto=tipo_producto,
                                motivo="Selección Manual"
                            ))
                        
                        HistorialEliminacion.objects.bulk_create(historial_objs)
                        
                        count, _ = items_a_borrar.delete()
                        messages.success(request, f'Se han eliminado {count} producto(s) correctamente.')
                    
                    return redirect('eliminar')
                except Exception as e:
                    messages.error(request, f'Error al eliminar productos: {str(e)}')
                    return redirect('eliminar')
        
        # --- CASO 3: ELIMINAR TODOS ---
        elif 'eliminar_todos' in request.POST:
            tipo_producto = request.POST.get('tipo_producto')
            try:
                model_class = None
                if tipo_producto == 'Tronco': model_class = Tronco
                elif tipo_producto == 'Piernas': model_class = Piernas
                elif tipo_producto == 'Zapatos': model_class = Zapatos
                elif tipo_producto == 'Complemento': model_class = Complemento
                
                if model_class:
                    items_a_borrar = model_class.objects.all()
                    
                    historial_objs = [
                        HistorialEliminacion(
                            usuario=request.user,
                            producto_modelo=item.modelo,
                            marca=item.marca.nombre if item.marca else "Sin Marca",
                            tipo_producto=tipo_producto,
                            motivo="Vaciado de Categoría Completa"
                        ) for item in items_a_borrar
                    ]
                    HistorialEliminacion.objects.bulk_create(historial_objs)
                    
                    count, _ = items_a_borrar.delete()
                    messages.success(request, f'Se han eliminado {count} producto(s).')
                
                return redirect('eliminar')
            except Exception as e:
                messages.error(request, f'Error al eliminar todos los productos: {str(e)}')
                return redirect('eliminar')
    
    return render(request, "producto/eliminar.html", {
        'productos': productos,
        'tipo_producto': tipo_producto,
        'historial_bajas': historial_bajas # Variable clave para la tabla
    })

@login_required
def Admin_User(request):
    # Usamos select_related para optimizar la consulta de preferencias
    users = User.objects.select_related('preferencias').all().order_by('id')
    tiendas = Tienda.objects.all()
    
    if request.method == 'POST':
        if 'delete_user' in request.POST:
            user_id = request.POST.get('user_id')
            try:
                user = User.objects.get(id=user_id)
                user.delete()
                messages.success(request, f'Usuario {user.username} eliminado correctamente.')
            except User.DoesNotExist:
                messages.error(request, 'El usuario no existe.')
            return redirect('Admin_User')
            
        elif 'update_user' in request.POST:
            user_id = request.POST.get('user_id')
            username = request.POST.get(f'username_{user_id}')
            email = request.POST.get(f'email_{user_id}')
            first_name = request.POST.get(f'first_name_{user_id}')
            last_name = request.POST.get(f'last_name_{user_id}')
            is_staff = request.POST.get(f'is_staff_{user_id}') == 'on'
            tienda_id = request.POST.get(f'tienda_{user_id}')
            
            try:
                user = User.objects.get(id=user_id)
                user.username = username
                user.email = email
                user.first_name = first_name
                user.last_name = last_name
                user.is_staff = is_staff
                user.save()
                
                # Lógica para asignar Tienda en PreferenciasUsuario
                # get_or_create asegura que no falle si el usuario no tenía preferencias previas
                prefs, created = PreferenciasUsuario.objects.get_or_create(user=user)
                
                if tienda_id and tienda_id != "":
                    try:
                        tienda_obj = Tienda.objects.get(id=tienda_id)
                        prefs.tienda_asociada = tienda_obj
                    except Tienda.DoesNotExist:
                        prefs.tienda_asociada = None
                else:
                    # Si se selecciona la opción vacía, se desvincula la tienda
                    prefs.tienda_asociada = None
                
                prefs.save()
                
                messages.success(request, f'Usuario {user.username} actualizado correctamente.')
            except User.DoesNotExist:
                messages.error(request, 'El usuario no existe.')
            return redirect('Admin_User')
    
    return render(request, "Usuarios/Admin_User.html", {
        'users': users,
        'tiendas': tiendas
    })

def IniciarSesion(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            return redirect('index')  # Redirige a la página principal después del login
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')
    
    return render(request, "Usuarios/IniciarSesion.html")

def cerrar_sesion(request):
    logout(request)
    return redirect('index')

def Registrarse(request):
    if request.method == 'POST':
        form = UserCreationForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.first_name = request.POST.get('first_name')
            user.last_name = request.POST.get('last_name')
            user.email = request.POST.get('email')
            user.save()

            recibir_notificaciones = request.POST.get('recibir_notificaciones') == 'on'
            PreferenciasUsuario.objects.create(user=user, recibir_notificaciones=recibir_notificaciones)

            messages.success(request, f'¡Cuenta creada para {user.username}!')
            return redirect('IniciarSesion')
    else:
        form = UserCreationForm()

    # 🔹 Aplica clases de Bootstrap a todos los campos
    for field in form.fields.values():
        field.widget.attrs['class'] = 'form-control'

    return render(request, "Usuarios/Registrarse.html", {'form': form})


def QuienesSomos(request):
    return render(request, "Footer/QuienesSomos.html")

def SeSocio(request):
    return render(request, "Footer/SeSocio.html")

def TrabajaConNosotros(request):
    if request.method == 'POST':
        form = TrabajaConNosotrosForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Tu solicitud ha sido enviada con éxito! Nos pondremos en contacto contigo pronto.')
            return redirect('TrabajaConNosotros')
    else:
        form = TrabajaConNosotrosForm()
    
    return render(request, "Footer/TrabajaConNosotros.html", {'form': form})

def TerminosCondiciones(request):
    return render(request, "Footer/TerminosCondiciones.html")

def SeSocio(request):
    if request.method == 'POST':
        form = SeSocioForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, '¡Tu solicitud ha sido enviada con éxito! Nos pondremos en contacto contigo pronto.')
            return redirect('SeSocio')
    else:
        form = SeSocioForm()
    
    return render(request, "Footer/SeSocio.html", {'form': form})

def commit_pay(request):
    return render(request, "Ventas/commit_pay.html")

def pago_error(request):
    return render(request, "Ventas/pago_error.html")

def send_pay(request):
    return render(request, "Ventas/send_pay.html")

def exportar_envios_no_procesados(request):
    # Obtener todos los envíos no procesados
    envios = EnvioModel.objects.filter(procesado=False).order_by('fecha')
    
    # Crear un libro de Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Envios no procesados"
    
    # Encabezados
    headers = [
        'ID', 'Nombre', 'Email', 'Teléfono', 'Dirección', 
        'Ciudad', 'Región', 'Código Postal', 'Courier', 
        'Subtotal', 'Costo Envío', 'Total', 'Marcas', 'Fecha'
    ]
    
    # Estilos para encabezados
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    alignment = Alignment(horizontal='center', vertical='center')
    
    # Escribir encabezados
    for col_num, header in enumerate(headers, 1):
        col_letter = get_column_letter(col_num)
        cell = ws[f"{col_letter}1"]
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        ws.column_dimensions[col_letter].width = 20
    
    # Escribir datos
    for row_num, envio in enumerate(envios, 2):
        ws.cell(row=row_num, column=1, value=envio.id)
        ws.cell(row=row_num, column=2, value=envio.nombre)
        ws.cell(row=row_num, column=3, value=envio.email)
        ws.cell(row=row_num, column=4, value=envio.telefono)
        ws.cell(row=row_num, column=5, value=envio.direccion)
        ws.cell(row=row_num, column=6, value=envio.ciudad)
        ws.cell(row=row_num, column=7, value=envio.get_region_display())
        ws.cell(row=row_num, column=8, value=envio.codigo_postal)
        ws.cell(row=row_num, column=9, value=envio.get_courier_display())
        ws.cell(row=row_num, column=10, value=envio.monto_carrito)
        ws.cell(row=row_num, column=11, value=envio.costo_envio)
        ws.cell(row=row_num, column=12, value=envio.total_pagar)
        ws.cell(row=row_num, column=13, value=envio.marcas_compradas)
        ws.cell(row=row_num, column=14, value=envio.fecha.strftime("%d/%m/%Y %H:%M"))
    
    # Preparar la respuesta
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="envios_no_procesados.xlsx"'
    
    # Guardar el libro en la respuesta
    wb.save(response)
    
    return response